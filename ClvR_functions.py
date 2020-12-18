"""
Functions used for modeling in the Hay lab's 2lClvR publication.
"""

import copy
import itertools as itt
import numpy as np
from openpyxl import load_workbook
import os
import pandas as pd
import time

import matplotlib.pyplot as plt
import holoviews as hv

import bebi103

hv.extension('bokeh')


def allele_list(genotype, list_type, p1 = 0):
    """allele_list generates a single list of alleles from the input 
    genotype. 
    
    list_type determines whether output is as a genotype or 
    haplotypes. p1 determines whether the mother(0) or the father (1)
    is the primary haplotype of interest.
    
    genotype_list, simple list of all alleles for specified 
        genotype
    mother_haplotype, list of alleles from the mother
    father_haplotype, list of alleles from the father
    """
    
    # For a genotype, just pool all alleles into a single list
    if list_type == 'genotype':

        genotype_list = []

        for locus_index, locus in enumerate(genotype):
            for allele_index, allele in enumerate(locus):
                genotype_list.append(genotype[locus_index][allele_index])

        return(genotype_list)

    # For a haplotype, pool alleles from mother and father separately
    elif list_type == 'haplotype':

        mother_haplotype = []
        father_haplotype = []

        for locus_index, locus in enumerate(genotype):
            mother_haplotype.append(genotype[locus_index][0])
            father_haplotype.append(genotype[locus_index][1])

        if p1 == 0:
            return(mother_haplotype, father_haplotype)
        
        elif p1 == 1:
            return(father_haplotype, mother_haplotype)


def all_option(subset, overset, replacement):
    """all_option checks to make sure that the overset contains 
    everything within the subset, with or without replacement. 
    Couldn't find a function that did this natively, if Justin reads 
    this and knows of one please replace it!
    """
    subsetcopy = subset.copy()
    oversetcopy = overset.copy()
    
    if replacement:
        check = 1
        for item in subsetcopy:
            if not item in oversetcopy:
                check = 0
                break
    
        return(check)
    
    else:
        check = 1
        for item in subsetcopy:
            if item in oversetcopy:
                oversetcopy.remove(item)
                
            else:
                check = 0
                break
    
        return(check)


def drive_fitness(lethal, rescue, genotype_list):
    """drive_fitness makes a list of fitness affects drive_f_c 
    based on special, genetic lethality and rescue conditions 
    for a given genotype_list. 
    
    lethal is a list of lists, where each inner list contains 
    the allele(s) that constitute a lethal condition and the 
    outer list contains all such lethal condition lists.
    rescue is a list of lists of lists, where the middle lists 
    contain all possible rescue conditions for the 
    corresponding lethal condition, the innermost lists contain
    all allele(s) necessary for that specific rescue condition, 
    and the outer list contains all such sets of rescue 
    condtions for each lethal condition.
    genotype_list is the list of lists of all genotypes for 
    either males or females, which is used to determine which 
    genotypes succumb to the lethality and which are rescued.
    """
    
    drive_f_c = [None]*len(genotype_list)
    
    for allele_list_ind, allele_list in enumerate(genotype_list):
        fitness_cost = 0
        fitness_cost_per_condtion = 0
        
        # for each lethal element
        for element_ind, element in enumerate(lethal):
            
            # check for lethality
            if all_option(element, allele_list, 0):
                
                # for each rescue condition
                for rescue_condition in rescue[element_ind]:
                    
                    # check for rescue
                    if all_option(rescue_condition, allele_list, 0):
                        fitness_cost_per_condtion = 0
                        break
                        
                    else:
                        fitness_cost_per_condtion = 1
                fitness_cost += fitness_cost_per_condtion
        if fitness_cost > 1.1:
            fitness_cost = 1
        drive_f_c[allele_list_ind] = fitness_cost
    
    return(np.array(drive_f_c))


def collate_f_c(alleles_f_c, drive_f_c, genotype_list, gender_alleles, f_c_type):
    """collate_f_c takes in a list of fitness costs for each allele and 
    drive specific lethalities and generates a final list of fitness
    costs for a given gender.
    """
    
    f_c = [0]*len(drive_f_c)
    
    if f_c_type == 'dominant':
        for geno_ind, genotype in enumerate(genotype_list):
            for allele_ind, allele in enumerate(gender_alleles):
                added_f_c = (allele in genotype)*alleles_f_c[allele_ind]

                if f_c[geno_ind] + drive_f_c[geno_ind] + added_f_c >= 1:
                    f_c[geno_ind] = 1
                    
                else:
                    f_c[geno_ind] += drive_f_c[geno_ind] + added_f_c
                    
    elif f_c_type == 'additive':
        for geno_ind, genotype in enumerate(genotype_list):
            for allele_ind, allele in enumerate(gender_alleles):
                added_f_c = genotype.count(allele)*alleles_f_c[allele_ind]

                if f_c[geno_ind] + drive_f_c[geno_ind] + added_f_c >= 1:
                    f_c[geno_ind] = 1
                    
                else:
                    f_c[geno_ind] += drive_f_c[geno_ind] + added_f_c
                    
    return(f_c)


def matrix_evaluator(male_cross_matrix, female_cross_matrix, d_a, 
                     d_a_males_string, d_a_females_string, 
                     d_a_males_rates_string, d_a_females_rates_string, 
                     num_loci, num_male_geno, num_female_geno, 
                     drive_activites_total_counter, 
                     m_li_b=0, f_li_b=0, m_li_inv_b=0, f_li_inv_b=0):
    """matrix_evaluator a deterministic, discrete generation, population 
    frequency dynamics model that takes in the cross_matrices and the user 
    input drive parameters to evaulate the specified gene drive's behavior 
    over time.
    
    
    For all other parameters, see drive_data functions.
    """
    
    if len(d_a[0]) == drive_activites_total_counter:
        if len(d_a) == 1:
            d_a += d_a
            
        if len(d_a) == 2:
            # drive activity for males or females (mxn table, m = number of 
            # genotypes, n = number of drive activities)
            d_a_m = []  
            d_a_f = []
            
            # drive activity for males or females rates (1xn list, n = drive 
            # activity rates for Males)
            d_a_m_r = []
            d_a_f_r = []

            for activity in range(drive_activites_total_counter):
                d_a_m.append([])
                d_a_f.append([])

                d_a_m_r.append([])
                d_a_f_r.append([])

                for male in range(num_male_geno):
                    d_a_m[activity].append([])

                for female in range(num_female_geno):
                    d_a_f[activity].append([])

            for activity in range(drive_activites_total_counter):
                d_a_m_r[activity] = eval(d_a_males_rates_string[activity])
                d_a_f_r[activity] = eval(d_a_females_rates_string[activity])

                for male in range(num_male_geno):
                    d_a_m[activity][male] = eval(d_a_males_string[activity][male])

                for female in range(num_female_geno):
                    d_a_f[activity][female] = eval(d_a_females_string[activity][female])
        
        else:
            RuntimeError(f'D_A must be of length 1 or 2.')
    
    elif len(d_a[0][0]) == num_male_geno:
        if len(d_a) == 1 and num_male_geno == num_female_geno:
            d_a_m = d_a[0]
            d_a_f = d_a[0]
        
        elif len(d_a) == 2 and len(d_a[1][0]) == num_female_geno:
            d_a_m = d_a[0]
            d_a_f = d_a[1]
        
        else:
            RuntimeError(f'd_a isn\'t of length 1 or 2, len(d_a[0][0]) \
                != num_male_geno, and/or len(d_a[1][0]) != num_female_geno')
    
    # Set linkage rates (aka recombination distance) between loci
    if m_li_b == 0 and f_li_b == 0:
        m_li = np.ones(num_loci-1)*0.5
        f_li = np.ones(num_loci-1)*0.5
    else:
        m_li = np.array(m_li_b)*0.5+0.5
        f_li = np.array(f_li_b)*0.5+0.5
        
    m_un = 1-m_li
    f_un = 1-f_li
    
    if m_li_inv_b == 0 and f_li_inv_b == 0:
        m_li_inv = np.ones(num_loci-1)*0.5
        f_li_inv = np.ones(num_loci-1)*0.5
    else:
        m_li_inv = np.array(m_li_inv_b)*0.5+0.5
        f_li_inv = np.array(f_li_inv_b)*0.5+0.5
        
    m_un_inv = 1-m_li_inv
    f_un_inv = 1-f_li_inv

    # Convert cross matrices to numerical form
    male_cm = np.zeros([num_male_geno, num_male_geno, num_female_geno])
    female_cm = np.zeros([num_female_geno, num_male_geno, num_female_geno])

    for ind1, genotype in enumerate(male_cross_matrix):
        for ind2, mother in enumerate(genotype):
            for ind3, father in enumerate(mother):
                if isinstance(male_cross_matrix[ind1][ind2][ind3], str):
                    male_cm[ind1][ind2][ind3] = eval(male_cross_matrix[ind1][ind2][ind3])
                elif male_cross_matrix[ind1][ind2][ind3] == None:
                    male_cm[ind1][ind2][ind3] = 0
                else:
                    print('Inappropriate cross matrix entry')

    for ind1, genotype in enumerate(female_cross_matrix):
        for ind2, mother in enumerate(genotype):
            for ind3, father in enumerate(mother):
                if isinstance(female_cross_matrix[ind1][ind2][ind3], str):
                    female_cm[ind1][ind2][ind3] = eval(female_cross_matrix[ind1][ind2][ind3])
                elif female_cross_matrix[ind1][ind2][ind3] == None:
                    female_cm[ind1][ind2][ind3] = 0
                else:
                    print('Inappropriate cross matrix entry')
    return(male_cm, female_cm)


def drive_data_1pop(master_list, d_a, lethal, rescue, 
                    m_eval = 0, num_gens=600, 
                    i_f=[0.25, 0.25], i_g=[0, 0], 
                    i_f_inv=[0, 0], i_inv_g=[203,283], 
                    f_c=[0.0, 0.0], f_c_type='additive', 
                    m_li_b=0, f_li_b=0, m_li_inv_b=0, f_li_inv_b=0, 
                    a_r=0, a_r_i_f=[0.25, 0.25], a_r_i_g=[-1, -1], 
                    a_r_g=50, a_r_f=50):
    """drive_data_1pop takes in cross matrices and all desired 
    run-specific parameters, performs a deterministic, population 
    frequency, discrete-generation population dynamic simulation, 
    and returns phenotype and allele frequencies for the specified 
    number of generations of simulation.
    
    d_a = drive activity, a list of lists of the rates of all drive 
    activities, the first listcontaining rates for males, the second
    containing rates for females. Must be of one of the three 
    following forms: two item list where the first index is a list of 
    lists whose dimesnions are the number of drive activies by the 
    number of male genotypes and the second list is the same but for 
    female genotypes, a two item list where the first list is the 
    number of male drive activities long and the second is the number of 
    female drive activities long, or a list whose length is the number
    of gene drive activites for either males or females, where all 
    entries are a number between 0 and 1.
    
    lethal is a list of lists, where each inner list contains 
    the allele(s) that constitute a lethal condition and the 
    outer list contains all such lethal condition lists.
    
    rescue is a list of lists of lists, where the middle lists 
    contain all possible rescue conditions for the 
    corresponding lethal condition, the innermost lists contain
    all allele(s) necessary for that specific rescue condition, 
    and the outer list contains all such sets of rescue 
    condtions for each lethal condition.
    
    m_eval = matrix evaluated, either 0 or a list containing the two 
    evaluated cross matrices.
    
    num_gens, number of generations the simulation runs for.
    i_f, introduction frequency(ies) for the introduced male
    and female genotypes, respectively.
    
    i_g, the introduction genotypes for males and females, 
    respectively.
    
    f_c = fitness cost, two item list of either one float per item 
    (fitness cost applied to first allele of first locus), the two 
    items are lists whose lengths are the number of male alleles and 
    number of female alleles, respectively, or the two items are lists
    whose lengths are the number of male and female genotypes, 
    respectively. All values must be between 0 and 1.
    
    f_c_type, determines whether the fitness costs are additive or 
    dominant.
    
    m_li_b and f_li_b are the male and female linkage rates, which 
    represent the probability that two alleles from two different loci 
    segregate together based on the recombination distance between them.
    Must be a value between 0 and 1. Can be left empty or their lengths
    must be equal to the number of loci-1.
    
    a_r = additional releases, determines whether or not there will be 
    additional releases later in the simualtion.
    
    a_r_i_f, the introduction frequencies for additional releases. 
    Follow the same rules as i_f.
    
    a_r_i_g, the genotypes for additional releases. Follow same rules as
    i_g.
    
    a_r_g, generation at which first additional release occurs.
    
    a_r_f, the number of generations between each successive 
    additional release.
    """
    num_loci = master_list[0]
    drive_activites_total_counter = master_list[1]
    num_male_geno = master_list[2]
    num_female_geno = master_list[3]
    males_g_a = master_list[4]
    females_g_a = master_list[5]
    male_alleles = master_list[6]
    female_alleles = master_list[7]
    all_alleles = master_list[8]
    male_a_freq_string = master_list[9]
    female_a_freq_string = master_list[10]
    male_p_freq_string = master_list[11]
    female_p_freq_string = master_list[12]
    all_a_freq_string = master_list[13]
    all_p_freq_string = master_list[14]
    d_a_males_string = master_list[15]
    d_a_females_string = master_list[16]
    d_a_males_rates_string = master_list[17]
    d_a_females_rates_string = master_list[18]
    male_cross_matrix = master_list[19]
    female_cross_matrix = master_list[20]
    
    # Initialize actual male and female population proportions
    M = np.matrix(np.zeros((num_gens+1, num_male_geno)))
    F = np.matrix(np.zeros((num_gens+1, num_female_geno)))
    
    # Set initial population proportions for male (M1) and female (F1)
    if isinstance(i_f[0], float) and isinstance(i_f[1], float):
        M1 = np.zeros(num_male_geno)
        M1[i_g[0]] = i_f[0]
        M1[-1] = (1-np.sum(i_f)-np.sum(i_f_inv))/2

        F1 = np.zeros(num_female_geno)
        F1[i_g[1]] = i_f[1]
        F1[-1] = (1-np.sum(i_f)-np.sum(i_f_inv))/2
        
        if np.sum(i_f_inv) != 0:
            x = len(i_inv_g)
            for genotype in i_inv_g:
                M1[genotype] = i_f_inv[0]/x
                F1[genotype] = i_f_inv[1]/x
        
    elif len(i_f[0]) == num_male_geno and len(i_f[1]) == num_female_geno:
        M1 = i_f[0]
        F1 = i_f[1]

    else:
        raise RuntimeError(f'IF must be a two item list or a list of two \
            lists of length num_male_geno and num_female_geno, respectively')

    M[0,:] = M1
    F[0,:] = F1
    
    if isinstance(f_c[0], float) and isinstance(f_c[1], float):
        males_a_f_c = np.zeros(len(male_alleles))
        females_a_f_c = np.zeros(len(female_alleles))
        males_a_f_c[0] = f_c[0]
        females_a_f_c[0] = f_c[1]
        males_f_c = collate_f_c(males_a_f_c, drive_fitness(lethal, rescue, males_g_a), males_g_a, male_alleles, f_c_type)
        females_f_c = collate_f_c(females_a_f_c, 
                                drive_fitness(lethal, rescue, females_g_a), females_g_a, female_alleles, f_c_type)
    
    elif len(f_c[0]) == len(male_alleles) and len(f_c[1]) == len(female_alleles):
        males_a_f_c = f_c[0]
        females_a_f_c = f_c[1]
        males_f_c = collate_f_c(males_a_f_c, drive_fitness(lethal, rescue, males_g_a), males_g_a, male_alleles, f_c_type)
        females_f_c = collate_f_c(females_a_f_c, 
                                drive_fitness(lethal, rescue, females_g_a), females_g_a, female_alleles, f_c_type)

    elif len(f_c[0]) == num_male_geno and len(f_c[1]) == num_female_geno:
        males_f_c = f_c[0]
        females_f_c = f_c[1]

    else:
        raise RuntimeError(f'FC must be a two item list of lists of lengths len(male_alleles) and len(female_alleles) ot lengths num_male_geno and num_female_geno, respectively')
    
    for genotype_ind, genotype in enumerate(males_g_a):
        if genotype[0] == 'I' and genotype[2] != 'J':
            males_f_c[genotype_ind] = 1

        if genotype[1] == 'I' and genotype[3] != 'J':
            males_f_c[genotype_ind] = 1

        if genotype[2] == 'J' and genotype[0] != 'I':
            males_f_c[genotype_ind] = 1

        if genotype[3] == 'J' and genotype[1] != 'I':
            males_f_c[genotype_ind] = 1
    
    for genotype_ind, genotype in enumerate(females_g_a):
        if genotype[0] == 'I' and genotype[2] != 'J':
            females_f_c[genotype_ind] = 1

        if genotype[1] == 'I' and genotype[3] != 'J':
            females_f_c[genotype_ind] = 1

        if genotype[2] == 'J' and genotype[0] != 'I':
            females_f_c[genotype_ind] = 1

        if genotype[3] == 'J' and genotype[1] != 'I':
            females_f_c[genotype_ind] = 1
    
    if m_eval == 0:
        male_cm, female_cm = matrix_evaluator(male_cross_matrix, female_cross_matrix, d_a, 
                                              d_a_males_string, d_a_females_string,
                                              d_a_males_rates_string, d_a_females_rates_string,
                                              num_loci, num_male_geno, num_female_geno, 
                                              drive_activites_total_counter, 
                                              m_li_b, f_li_b, m_li_inv_b, f_li_inv_b)
        
    else:
        male_cm = m_eval[0]
        female_cm = m_eval[1]
    
    # Initialize temporary current population frequencies
    M_temp = np.array(np.zeros(num_male_geno))
    F_temp = np.array(np.zeros(num_female_geno))

    # Simulate population
    a_r_counter = 0
    sigma = np.zeros(num_gens+1)
    
    for gen in range(1,num_gens+1):

        # Generate proportion of each genotype pairing
        present_cross = np.transpose(M[gen-1])*F[gen-1]

        # Generate gross proportions for each genotype
        for male_geno in range(num_male_geno):
            M_temp[male_geno] = np.sum(np.multiply(male_cm[male_geno][:][:], present_cross))

        for female_geno in range(num_female_geno):
            F_temp[female_geno] = np.sum(np.multiply(female_cm[female_geno][:][:], present_cross))

        sigma[gen] = np.sum(np.multiply(M_temp, np.subtract([1]*len(males_f_c), males_f_c))) \
            + np.sum(np.multiply(F_temp, np.subtract([1]*len(females_f_c), females_f_c)))

        if gen%(a_r_g + a_r_f*a_r_counter) == False and a_r_counter < a_r:
            M[gen][:] = np.multiply(np.multiply(M_temp, np.subtract([1]*len(males_f_c), males_f_c))/sigma[gen], 1-np.sum(a_r_i_f))
            F[gen][:] = np.multiply(np.multiply(F_temp, np.subtract([1]*len(females_f_c), females_f_c))/sigma[gen], 1-np.sum(a_r_i_f))
            M[gen, a_r_i_g[0]] += a_r_i_f[0]
            F[gen, a_r_i_g[1]] += a_r_i_f[1]
            a_r_counter += 1

        else:
            M[gen][:] = np.multiply(M_temp, np.subtract([1]*len(males_f_c), males_f_c))/sigma[gen]
            F[gen][:] = np.multiply(F_temp, np.subtract([1]*len(females_f_c), females_f_c))/sigma[gen]
    
    # Reform data from genotype frequencies to 
    # allele/'phenotype' frequencies
    male_a_freqs = np.zeros([num_gens+1,len(male_a_freq_string)])
    female_a_freqs = np.zeros([num_gens+1,len(female_a_freq_string)])
    all_a_freqs = np.zeros([num_gens+1,len(all_a_freq_string)])
    male_p_freqs = np.zeros([num_gens+1,len(male_p_freq_string)])
    female_p_freqs = np.zeros([num_gens+1,len(female_p_freq_string)])
    all_p_freqs = np.zeros([num_gens+1,len(all_p_freq_string)])

    for f_i, frequency in enumerate(male_a_freq_string):
        for g in range(num_gens+1):
            male_a_freqs[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(male_p_freq_string):
        for g in range(num_gens+1):
            male_p_freqs[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(female_a_freq_string):
        for g in range(num_gens+1):
            female_a_freqs[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(female_p_freq_string):
        for g in range(num_gens+1):
            female_p_freqs[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(all_a_freq_string):
        for g in range(num_gens+1):
            all_a_freqs[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(all_p_freq_string):
        for g in range(num_gens+1):
            all_p_freqs[g][f_i] = eval(frequency)
    
    return(all_p_freqs, all_a_freqs)


def drive_data_2pop(master_list, d_a, lethal, rescue, 
                    m_eval = 0, num_gens=600, 
                    i_f=[0.25, 0.25], i_g=[0, 0], 
                    i_f_inv=[0, 0], i_inv_g=[203,283], 
                    f_c=[0.0, 0.0], f_c_type='additive', 
                    m_li_b=0, f_li_b=0, m_li_inv_b=0, f_li_inv_b=0, 
                    m_r=0.01, 
                    a_r=0, a_r_i_f=[0.25, 0.25], a_r_i_g=[-1, -1], 
                    a_r_g=50, a_r_f=50):
    """drive_data_2pop performs the same kind of simulation as 
    drive_data_1pop but simulates two populations that mix via migration.
    
    m_r, the migration rate between two populations. If its a 
    single value there is equal migration in both directions, if there
    are two values the first is the migration rate from pop1 to pop2 and
    the second is vice versa. Value(s) must be between 0 and 1, expected 
    to be <0.1.
    For all other parameters, see drive_data functions.
    """
    num_loci = master_list[0]
    drive_activites_total_counter = master_list[1]
    num_male_geno = master_list[2]
    num_female_geno = master_list[3]
    males_g_a = master_list[4]
    females_g_a = master_list[5]
    male_alleles = master_list[6]
    female_alleles = master_list[7]
    all_alleles = master_list[8]
    male_a_freq_string = master_list[9]
    female_a_freq_string = master_list[10]
    male_p_freq_string = master_list[11]
    female_p_freq_string = master_list[12]
    all_a_freq_string = master_list[13]
    all_p_freq_string = master_list[14]
    d_a_males_string = master_list[15]
    d_a_females_string = master_list[16]
    d_a_males_rates_string = master_list[17]
    d_a_females_rates_string = master_list[18]
    male_cross_matrix = master_list[19]
    female_cross_matrix = master_list[20]

    # Initialize actual male and female population proportions
    M_1 = np.matrix(np.zeros((num_gens+1, num_male_geno)))
    F_1 = np.matrix(np.zeros((num_gens+1, num_female_geno)))
    
    M_2 = np.matrix(np.zeros((num_gens+1, num_male_geno)))
    F_2 = np.matrix(np.zeros((num_gens+1, num_female_geno)))
    
    # Set initial population proportions for male (M1) and female (F1)
    if isinstance(i_f[0], float) and isinstance(i_f[1], float):
        M1 = np.zeros(num_male_geno)
        M1[i_g[0]] = i_f[0]
        M1[-1] = (1-np.sum(i_f)-np.sum(i_f_inv))/2

        F1 = np.zeros(num_female_geno)
        F1[i_g[1]] = i_f[1]
        F1[-1] = (1-np.sum(i_f)-np.sum(i_f_inv))/2
        
        if np.sum(i_f_inv) != 0:
            x = len(i_inv_g)
            for genotype in i_inv_g:
                M1[genotype] = i_f_inv[0]/x
                F1[genotype] = i_f_inv[1]/x
        
    elif len(i_f[0]) == num_male_geno and len(i_f[1]) == num_female_geno:
        M1 = i_f[0]
        F1 = i_f[1]

    else:
        raise RuntimeError(f'IF must be a two item list or a list of two \
            lists of length num_male_geno and num_female_geno, respectively')

    M_1[0,:] = M1
    F_1[0,:] = F1
    
    M_2[0,-1] = 1/2
    F_2[0,-1] = 1/2
    
    if isinstance(f_c[0], float) and isinstance(f_c[1], float):
        males_a_f_c = np.zeros(len(male_alleles))
        females_a_f_c = np.zeros(len(female_alleles))
        males_a_f_c[0] = f_c[0]
        females_a_f_c[0] = f_c[1]
        males_f_c = collate_f_c(males_a_f_c, drive_fitness(lethal, rescue, males_g_a), males_g_a, male_alleles, f_c_type)
        females_f_c = collate_f_c(females_a_f_c, 
                                drive_fitness(lethal, rescue, females_g_a), females_g_a, female_alleles, f_c_type)
    
    elif len(f_c[0]) == len(male_alleles) and len(f_c[1]) == len(female_alleles):
        males_a_f_c = f_c[0]
        females_a_f_c = f_c[1]
        males_f_c = collate_f_c(males_a_f_c, drive_fitness(lethal, rescue, males_g_a), males_g_a, male_alleles, f_c_type)
        females_f_c = collate_f_c(females_a_f_c, 
                                drive_fitness(lethal, rescue, females_g_a), females_g_a, female_alleles, f_c_type)

    elif len(f_c[0]) == num_male_geno and len(f_c[1]) == num_female_geno:
        males_f_c = f_c[0]
        females_f_c = f_c[1]

    else:
        raise RuntimeError(f'FC must be a two item list of lists of lengths \
            len(male_alleles) and len(female_alleles) ot lengths \
            num_male_geno and num_female_geno, respectively')
    
    for genotype_ind, genotype in enumerate(males_g_a):
        if genotype[0] == 'I' and genotype[2] != 'J':
            males_f_c[genotype_ind] = 1

#         elif genotype[2] == 'J':
#             males_f_c[genotype_ind] = 1

        if genotype[1] == 'I' and genotype[3] != 'J':
            males_f_c[genotype_ind] = 1

        if genotype[2] == 'J' and genotype[0] != 'I':
            males_f_c[genotype_ind] = 1

        if genotype[3] == 'J' and genotype[1] != 'I':
            males_f_c[genotype_ind] = 1
    
    for genotype_ind, genotype in enumerate(females_g_a):
        if genotype[0] == 'I' and genotype[2] != 'J':
            females_f_c[genotype_ind] = 1

#         elif genotype[2] == 'J':
#             females_f_c[genotype_ind] = 1

        if genotype[1] == 'I' and genotype[3] != 'J':
            females_f_c[genotype_ind] = 1

        if genotype[2] == 'J' and genotype[0] != 'I':
            females_f_c[genotype_ind] = 1

        if genotype[3] == 'J' and genotype[1] != 'I':
            females_f_c[genotype_ind] = 1
    
    
    
    if m_eval == 0:
        male_cm, female_cm = matrix_evaluator(male_cross_matrix, female_cross_matrix, d_a, 
                                              d_a_males_string, d_a_females_string,
                                              d_a_males_rates_string, d_a_females_rates_string,
                                              num_loci, num_male_geno, num_female_geno, 
                                              drive_activites_total_counter, 
                                              m_li_b, f_li_b, m_li_inv_b, f_li_inv_b)
        
    else:
        male_cm = m_eval[0]
        female_cm = m_eval[1]
    
    if isinstance(m_r, float):
        m_r = [m_r, m_r]
    
    elif len(m_r) == 2:
        pass
    
    else:
        raise RuntimeError(f'm_r must be a float or a list of two floats')
        
    # Initialize temporary current population frequencies
    M_1_temp = np.array(np.zeros(num_male_geno))
    F_1_temp = np.array(np.zeros(num_female_geno))

    M_2_temp = np.array(np.zeros(num_male_geno))
    F_2_temp = np.array(np.zeros(num_female_geno))


    # Simulate population
    a_r_counter = 0
    sigma_1 = np.zeros(num_gens+1)
    sigma_2 = np.zeros(num_gens+1)
    
    for gen in range(1,num_gens+1):

        # Generate proportion of each genotype pairing
        presentcross_1 = np.transpose(M_1[gen-1])*F_1[gen-1]
        presentcross_2 = np.transpose(M_2[gen-1])*F_2[gen-1]

        # Generate gross proportions for each genotype
        for male_geno in range(num_male_geno):
            M_1_temp[male_geno] = np.sum(np.multiply(male_cm[male_geno][:][:],presentcross_1))
            M_2_temp[male_geno] = np.sum(np.multiply(male_cm[male_geno][:][:],presentcross_2))

        for female_geno in range(num_female_geno):
            F_1_temp[female_geno] = np.sum(np.multiply(female_cm[female_geno][:][:],presentcross_1))
            F_2_temp[female_geno] = np.sum(np.multiply(female_cm[female_geno][:][:],presentcross_2))

        sigma_1[gen] = np.sum(np.multiply(M_1_temp, np.subtract([1]*len(males_f_c), males_f_c))) \
            + np.sum(np.multiply(F_1_temp, np.subtract([1]*len(females_f_c), females_f_c)))
        sigma_2[gen] = np.sum(np.multiply(M_2_temp, np.subtract([1]*len(males_f_c), males_f_c))) \
            + np.sum(np.multiply(F_2_temp, np.subtract([1]*len(females_f_c), females_f_c)))

        M_1[gen][:] = np.multiply(M_1_temp, np.subtract([1]*len(males_f_c), males_f_c))/sigma_1[gen]
        F_1[gen][:] = np.multiply(F_1_temp, np.subtract([1]*len(females_f_c), females_f_c))/sigma_1[gen]
        M_2[gen][:] = np.multiply(M_2_temp, np.subtract([1]*len(males_f_c), males_f_c))/sigma_2[gen]
        F_2[gen][:] = np.multiply(F_2_temp, np.subtract([1]*len(females_f_c), females_f_c))/sigma_2[gen]

        M_1_sub = np.multiply(M_1[gen][:], m_r[0])
        F_1_sub = np.multiply(F_1[gen][:], m_r[0])

        M_2_sub = np.multiply(M_2[gen][:], m_r[1])
        F_2_sub = np.multiply(F_2[gen][:], m_r[1])

        M_1[gen][:] = M_1[gen][:] + M_2_sub - M_1_sub
        F_1[gen][:] = F_1[gen][:] + F_2_sub - F_1_sub

        M_2[gen][:] = M_2[gen][:] + M_1_sub - M_2_sub
        F_2[gen][:] = F_2[gen][:] + F_1_sub - F_2_sub
        
        # It is unclear if this part is accurate or not, since we are 
        # nominally increasing the size of population 1, might need to 
        # tweak it or just assume that all the extra individuals die 
        # after a single generation and don't affect migration numbers
        if gen%(a_r_g + a_r_f*a_r_counter) == False and a_r_counter < a_r:
            M_1[gen] = np.multiply(M_1[gen][:], 1-np.sum(a_r_i_f))
            F_1[gen] = np.multiply(F_1[gen][:], 1-np.sum(a_r_i_f))
            M_1[gen, a_r_i_g[0]] += a_r_i_f[0]
            F_1[gen, a_r_i_g[0]] += a_r_i_f[1]
            a_r_counter += 1

    # Reform data from genotype frequencies to 
    # allele/'phenotype' frequencies for:
    # for pop 1
    M = M_1
    F = F_1
    
    male_a_freqs_1 = np.zeros([num_gens+1,len(male_a_freq_string)])
    female_a_freqs_1 = np.zeros([num_gens+1,len(female_a_freq_string)])
    all_a_freqs_1 = np.zeros([num_gens+1,len(all_a_freq_string)])
    male_p_freqs_1 = np.zeros([num_gens+1,len(male_p_freq_string)])
    female_p_freqs_1 = np.zeros([num_gens+1,len(female_p_freq_string)])
    all_p_freqs_1 = np.zeros([num_gens+1,len(all_p_freq_string)])

    for f_i, frequency in enumerate(male_a_freq_string):
        for g in range(num_gens+1):
            male_a_freqs_1[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(male_p_freq_string):
        for g in range(num_gens+1):
            male_p_freqs_1[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(female_a_freq_string):
        for g in range(num_gens+1):
            female_a_freqs_1[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(female_p_freq_string):
        for g in range(num_gens+1):
            female_p_freqs_1[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(all_a_freq_string):
        for g in range(num_gens+1):
            all_a_freqs_1[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(all_p_freq_string):
        for g in range(num_gens+1):
            all_p_freqs_1[g][f_i] = eval(frequency)

    # for pop 2
    M = M_2
    F = F_2

    male_a_freqs_2 = np.zeros([num_gens+1,len(male_a_freq_string)])
    female_a_freqs_2 = np.zeros([num_gens+1,len(female_a_freq_string)])
    all_a_freqs_2 = np.zeros([num_gens+1,len(all_a_freq_string)])
    male_p_freqs_2 = np.zeros([num_gens+1,len(male_p_freq_string)])
    female_p_freqs_2 = np.zeros([num_gens+1,len(female_p_freq_string)])
    all_p_freqs_2 = np.zeros([num_gens+1,len(all_p_freq_string)])
    M_wt_f_2 = np.zeros(num_gens+1)
    F_wt_f_2 = np.zeros(num_gens+1)

    for f_i, frequency in enumerate(male_a_freq_string):
        for g in range(num_gens+1):
            male_a_freqs_2[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(male_p_freq_string):
        for g in range(num_gens+1):
            male_p_freqs_2[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(female_a_freq_string):
        for g in range(num_gens+1):
            female_a_freqs_2[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(female_p_freq_string):
        for g in range(num_gens+1):
            female_p_freqs_2[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(all_a_freq_string):
        for g in range(num_gens+1):
            all_a_freqs_2[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(all_p_freq_string):
        for g in range(num_gens+1):
            all_p_freqs_2[g][f_i] = eval(frequency)
            
    return(all_p_freqs_1, all_a_freqs_1, all_p_freqs_2, all_a_freqs_2)


def drive_data_3pop(master_list, d_a, lethal, rescue, 
                    m_eval = 0, num_gens=600, 
                    i_f=[0.25, 0.25], i_g=[0, 0], 
                    i_f_inv=[0, 0], i_inv_g=[203,283], 
                    f_c=[0.0, 0.0], f_c_type='additive', 
                    m_li_b=0, f_li_b=0, m_li_inv_b=0, f_li_inv_b=0, 
                    m_r=0.01, 
                    a_r=0, a_r_i_f=[0.25, 0.25], a_r_i_g=[-1, -1], 
                    a_r_g=50, a_r_f=50):
    """drive_data_3pop performs the same kind of simulation as 
    drive_data_1pop but simulates three populations that mix via 
    migration sequentially (1<->2<->3).
    
    m_r, the migration rate between two populations. It can be a 
    single value if there is equal migration in both directions, or it 
    can be are four values where the first is the migration rate from 
    pop1 to pop2, the second is vice versa, and the third and fourth are
    the same but for the relation between pop2 and pop3. Value(s) must be 
    between 0 and 1, expected to be <0.1.
    For all other parameters, see drive_data functions.
    """
    num_loci = master_list[0]
    drive_activites_total_counter = master_list[1]
    num_male_geno = master_list[2]
    num_female_geno = master_list[3]
    males_g_a = master_list[4]
    females_g_a = master_list[5]
    male_alleles = master_list[6]
    female_alleles = master_list[7]
    all_alleles = master_list[8]
    male_a_freq_string = master_list[9]
    female_a_freq_string = master_list[10]
    male_p_freq_string = master_list[11]
    female_p_freq_string = master_list[12]
    all_a_freq_string = master_list[13]
    all_p_freq_string = master_list[14]
    d_a_males_string = master_list[15]
    d_a_females_string = master_list[16]
    d_a_males_rates_string = master_list[17]
    d_a_females_rates_string = master_list[18]
    male_cross_matrix = master_list[19]
    female_cross_matrix = master_list[20]

    # Initialize actual male and female population proportions
    M_1 = np.matrix(np.zeros((num_gens+1, num_male_geno)))
    F_1 = np.matrix(np.zeros((num_gens+1, num_female_geno)))
    
    M_2 = np.matrix(np.zeros((num_gens+1, num_male_geno)))
    F_2 = np.matrix(np.zeros((num_gens+1, num_female_geno)))
    
    M_3 = np.matrix(np.zeros((num_gens+1, num_male_geno)))
    F_3 = np.matrix(np.zeros((num_gens+1, num_female_geno)))
    
    # Set initial population proportions for male (M1) and female (F1)
    if isinstance(i_f[0], float) and isinstance(i_f[1], float):
        M1 = np.zeros(num_male_geno)
        M1[i_g[0]] = i_f[0]
        M1[-1] = (1-np.sum(i_f)-np.sum(i_f_inv))/2

        F1 = np.zeros(num_female_geno)
        F1[i_g[1]] = i_f[1]
        F1[-1] = (1-np.sum(i_f)-np.sum(i_f_inv))/2
        
        if np.sum(i_f_inv) != 0:
            x = len(i_inv_g)
            for genotype in i_inv_g:
                M1[genotype] = i_f_inv[0]/x
                F1[genotype] = i_f_inv[1]/x
        
    elif len(i_f[0]) == num_male_geno and len(i_f[1]) == num_female_geno:
        M1 = i_f[0]
        F1 = i_f[1]

    else:
        raise RuntimeError(f'IF must be a two item list or a list of two \
            lists of length num_male_geno and num_female_geno, respectively')

    M_1[0,:] = M1
    F_1[0,:] = F1
    
    M_2[0,-1] = 1/2
    F_2[0,-1] = 1/2
    
    M_3[0,-1] = 1/2
    F_3[0,-1] = 1/2
    
    if isinstance(f_c[0], float) and isinstance(f_c[1], float):
        males_a_f_c = np.zeros(len(male_alleles))
        females_a_f_c = np.zeros(len(female_alleles))
        males_a_f_c[0] = f_c[0]
        females_a_f_c[0] = f_c[1]
        males_f_c = collate_f_c(males_a_f_c, drive_fitness(lethal, rescue, males_g_a), males_g_a, male_alleles, f_c_type)
        females_f_c = collate_f_c(females_a_f_c, 
                                drive_fitness(lethal, rescue, females_g_a), females_g_a, female_alleles, f_c_type)
    
    elif len(f_c[0]) == len(male_alleles) and len(f_c[1]) == len(female_alleles):
        males_a_f_c = f_c[0]
        females_a_f_c = f_c[1]
        males_f_c = collate_f_c(males_a_f_c, drive_fitness(lethal, rescue, males_g_a), males_g_a, male_alleles, f_c_type)
        females_f_c = collate_f_c(females_a_f_c, 
                                drive_fitness(lethal, rescue, females_g_a), females_g_a, female_alleles, f_c_type)

    elif len(f_c[0]) == num_male_geno and len(f_c[1]) == num_female_geno:
        males_f_c = f_c[0]
        females_f_c = f_c[1]

    else:
        raise RuntimeError(f'FC must be a two item list of lists of lengths \
            len(male_alleles) and len(female_alleles) ot lengths \
            num_male_geno and num_female_geno, respectively')
    
    for genotype_ind, genotype in enumerate(males_g_a):
        if genotype[0] == 'I' and genotype[2] != 'J':
            males_f_c[genotype_ind] = 1

#         elif genotype[2] == 'J':
#             males_f_c[genotype_ind] = 1

        if genotype[1] == 'I' and genotype[3] != 'J':
            males_f_c[genotype_ind] = 1

        if genotype[2] == 'J' and genotype[0] != 'I':
            males_f_c[genotype_ind] = 1

        if genotype[3] == 'J' and genotype[1] != 'I':
            males_f_c[genotype_ind] = 1
    
    for genotype_ind, genotype in enumerate(females_g_a):
        if genotype[0] == 'I' and genotype[2] != 'J':
            females_f_c[genotype_ind] = 1

#         elif genotype[2] == 'J':
#             females_f_c[genotype_ind] = 1

        if genotype[1] == 'I' and genotype[3] != 'J':
            females_f_c[genotype_ind] = 1

        if genotype[2] == 'J' and genotype[0] != 'I':
            females_f_c[genotype_ind] = 1

        if genotype[3] == 'J' and genotype[1] != 'I':
            females_f_c[genotype_ind] = 1
    
    
    
    if m_eval == 0:
        male_cm, female_cm = matrix_evaluator(male_cross_matrix, female_cross_matrix, d_a, 
                                              d_a_males_string, d_a_females_string,
                                              d_a_males_rates_string, d_a_females_rates_string,
                                              num_loci, num_male_geno, num_female_geno, 
                                              drive_activites_total_counter, 
                                              m_li_b, f_li_b, m_li_inv_b, f_li_inv_b)
        
    else:
        male_cm = m_eval[0]
        female_cm = m_eval[1]
    
    if isinstance(m_r, float):
        m_r = [m_r, m_r, m_r, m_r]
    
    elif len(m_r) == 4:
        pass
    
    else:
        raise RuntimeError(f'm_r must be a float or a list of four floats')
        
    # Initialize temporary current population frequencies
    M_1_temp = np.array(np.zeros(num_male_geno))
    F_1_temp = np.array(np.zeros(num_female_geno))

    M_2_temp = np.array(np.zeros(num_male_geno))
    F_2_temp = np.array(np.zeros(num_female_geno))
    
    M_3_temp = np.array(np.zeros(num_male_geno))
    F_3_temp = np.array(np.zeros(num_female_geno))


    # Simulate population
    a_r_counter = 0
    sigma_1 = np.zeros(num_gens+1)
    sigma_2 = np.zeros(num_gens+1)
    sigma_3 = np.zeros(num_gens+1)
    
    for gen in range(1,num_gens+1):

        # Generate proportion of each genotype pairing
        presentcross_1 = np.transpose(M_1[gen-1])*F_1[gen-1]
        presentcross_2 = np.transpose(M_2[gen-1])*F_2[gen-1]
        presentcross_3 = np.transpose(M_3[gen-1])*F_3[gen-1]


        # Generate gross proportions for each genotype
        for male_geno in range(num_male_geno):
            M_1_temp[male_geno] = np.sum(np.multiply(male_cm[male_geno][:][:],presentcross_1))
            M_2_temp[male_geno] = np.sum(np.multiply(male_cm[male_geno][:][:],presentcross_2))
            M_3_temp[male_geno] = np.sum(np.multiply(male_cm[male_geno][:][:],presentcross_3))

        for female_geno in range(num_female_geno):
            F_1_temp[female_geno] = np.sum(np.multiply(female_cm[female_geno][:][:],presentcross_1))
            F_2_temp[female_geno] = np.sum(np.multiply(female_cm[female_geno][:][:],presentcross_2))
            F_3_temp[female_geno] = np.sum(np.multiply(female_cm[female_geno][:][:],presentcross_3))

        sigma_1[gen] = np.sum(np.multiply(M_1_temp, np.subtract([1]*len(males_f_c), males_f_c))) \
            + np.sum(np.multiply(F_1_temp, np.subtract([1]*len(females_f_c), females_f_c)))
        sigma_2[gen] = np.sum(np.multiply(M_2_temp, np.subtract([1]*len(males_f_c), males_f_c))) \
            + np.sum(np.multiply(F_2_temp, np.subtract([1]*len(females_f_c), females_f_c)))
        sigma_3[gen] = np.sum(np.multiply(M_3_temp, np.subtract([1]*len(males_f_c), males_f_c))) \
            + np.sum(np.multiply(F_3_temp, np.subtract([1]*len(females_f_c), females_f_c)))

        M_1[gen][:] = np.multiply(M_1_temp, np.subtract([1]*len(males_f_c), males_f_c))/sigma_1[gen]
        F_1[gen][:] = np.multiply(F_1_temp, np.subtract([1]*len(females_f_c), females_f_c))/sigma_1[gen]
        M_2[gen][:] = np.multiply(M_2_temp, np.subtract([1]*len(males_f_c), males_f_c))/sigma_2[gen]
        F_2[gen][:] = np.multiply(F_2_temp, np.subtract([1]*len(females_f_c), females_f_c))/sigma_2[gen]
        M_3[gen][:] = np.multiply(M_3_temp, np.subtract([1]*len(males_f_c), males_f_c))/sigma_3[gen]
        F_3[gen][:] = np.multiply(F_3_temp, np.subtract([1]*len(females_f_c), females_f_c))/sigma_3[gen]

        M_1_sub = np.multiply(M_1[gen][:], m_r[0])
        F_1_sub = np.multiply(F_1[gen][:], m_r[0])

        M_2_sub = np.multiply(M_2[gen][:], m_r[1])
        F_2_sub = np.multiply(F_2[gen][:], m_r[1])
        
        M_2_sub_2 = np.multiply(M_2[gen][:], m_r[2])
        F_2_sub_2 = np.multiply(F_2[gen][:], m_r[2])
        
        M_3_sub = np.multiply(M_3[gen][:], m_r[3])
        F_3_sub = np.multiply(F_3[gen][:], m_r[3])

        M_1[gen][:] = M_1[gen][:] + M_2_sub - M_1_sub
        F_1[gen][:] = F_1[gen][:] + F_2_sub - F_1_sub

        M_2[gen][:] = M_2[gen][:] + M_1_sub + M_3_sub - M_2_sub - M_2_sub_2
        F_2[gen][:] = F_2[gen][:] + F_1_sub + F_3_sub - F_2_sub - F_2_sub_2
        
        M_3[gen][:] = M_3[gen][:] + M_2_sub_2 - M_3_sub
        F_3[gen][:] = F_3[gen][:] + F_2_sub_2 - F_3_sub
        
        # It is unclear if this part is accurate or not, since we are 
        # nominally increasing the size of population 1, might need to 
        # tweak it or just assume that all the extra individuals die 
        # after a single generation and don't affect migration numbers
        if gen%(a_r_g + a_r_f*a_r_counter) == False and a_r_counter < a_r:
            M_1[gen] = np.multiply(M_1[gen][:], 1-np.sum(a_r_i_f))
            F_1[gen] = np.multiply(F_1[gen][:], 1-np.sum(a_r_i_f))
            M_1[gen, a_r_i_g[0]] += a_r_i_f[0]
            F_1[gen, a_r_i_g[0]] += a_r_i_f[1]
            a_r_counter += 1
    
    # Reform data from genotype frequencies to 
    # allele/'phenotype' frequencies for:
    # for pop 1
    M = M_1
    F = F_1

    male_a_freqs_1 = np.zeros([num_gens+1,len(male_a_freq_string)])
    female_a_freqs_1 = np.zeros([num_gens+1,len(female_a_freq_string)])
    all_a_freqs_1 = np.zeros([num_gens+1,len(all_a_freq_string)])
    male_p_freqs_1 = np.zeros([num_gens+1,len(male_p_freq_string)])
    female_p_freqs_1 = np.zeros([num_gens+1,len(female_p_freq_string)])
    all_p_freqs_1 = np.zeros([num_gens+1,len(all_p_freq_string)])

    for f_i, frequency in enumerate(male_a_freq_string):
        for g in range(num_gens+1):
            male_a_freqs_1[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(male_p_freq_string):
        for g in range(num_gens+1):
            male_p_freqs_1[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(female_a_freq_string):
        for g in range(num_gens+1):
            female_a_freqs_1[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(female_p_freq_string):
        for g in range(num_gens+1):
            female_p_freqs_1[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(all_a_freq_string):
        for g in range(num_gens+1):
            all_a_freqs_1[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(all_p_freq_string):
        for g in range(num_gens+1):
            all_p_freqs_1[g][f_i] = eval(frequency)

    # for pop 2
    M = M_2
    F = F_2

    male_a_freqs_2 = np.zeros([num_gens+1,len(male_a_freq_string)])
    female_a_freqs_2 = np.zeros([num_gens+1,len(female_a_freq_string)])
    all_a_freqs_2 = np.zeros([num_gens+1,len(all_a_freq_string)])
    male_p_freqs_2 = np.zeros([num_gens+1,len(male_p_freq_string)])
    female_p_freqs_2 = np.zeros([num_gens+1,len(female_p_freq_string)])
    all_p_freqs_2 = np.zeros([num_gens+1,len(all_p_freq_string)])

    for f_i, frequency in enumerate(male_a_freq_string):
        for g in range(num_gens+1):
            male_a_freqs_2[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(male_p_freq_string):
        for g in range(num_gens+1):
            male_p_freqs_2[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(female_a_freq_string):
        for g in range(num_gens+1):
            female_a_freqs_2[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(female_p_freq_string):
        for g in range(num_gens+1):
            female_p_freqs_2[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(all_a_freq_string):
        for g in range(num_gens+1):
            all_a_freqs_2[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(all_p_freq_string):
        for g in range(num_gens+1):
            all_p_freqs_2[g][f_i] = eval(frequency)
            
    
    # for pop 3
    M = M_3
    F = F_3

    male_a_freqs_3 = np.zeros([num_gens+1,len(male_a_freq_string)])
    female_a_freqs_3 = np.zeros([num_gens+1,len(female_a_freq_string)])
    all_a_freqs_3 = np.zeros([num_gens+1,len(all_a_freq_string)])
    male_p_freqs_3 = np.zeros([num_gens+1,len(male_p_freq_string)])
    female_p_freqs_3 = np.zeros([num_gens+1,len(female_p_freq_string)])
    all_p_freqs_3 = np.zeros([num_gens+1,len(all_p_freq_string)])

    for f_i, frequency in enumerate(male_a_freq_string):
        for g in range(num_gens+1):
            male_a_freqs_3[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(male_p_freq_string):
        for g in range(num_gens+1):
            male_p_freqs_3[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(female_a_freq_string):
        for g in range(num_gens+1):
            female_a_freqs_3[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(female_p_freq_string):
        for g in range(num_gens+1):
            female_p_freqs_3[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(all_a_freq_string):
        for g in range(num_gens+1):
            all_a_freqs_3[g][f_i] = eval(frequency)
    for f_i, frequency in enumerate(all_p_freq_string):
        for g in range(num_gens+1):
            all_p_freqs_3[g][f_i] = eval(frequency)
            
    return(all_p_freqs_1, all_a_freqs_1, all_p_freqs_2, all_a_freqs_2, all_p_freqs_3, all_a_freqs_3)


def append_df_to_excel(filename, sheet_name='Sheet1', 
                       df=None, pop_num=0, 
                       gene_drive_name=None, datatype=None, 
                       i_f=None, i_g=None, 
                       f_c_c=None, f_c_r=None, 
                       r_d=None, 
                       m_r=None, 
                       a_r=None, a_r_i_f=None, a_r_i_g=None,
                       a_r_g=None, a_r_f=None,
                       heatmap=0, allele=None, gens=None,
                       startcol=None, truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    # Written by MaxU on Stack Overflow, I adjusted to add by column 
    # instead of by row and the ability to add values pertaining to 
    # the run. Added values are  the data type (ie is it a phenotype
    # or an allele frequency), introduction frequency (IF), cas9 
    # fitness cost (cFC), rescue fitness cost (rFC, for when its not
    # in the same allele as cas9), recombination distance (r_d),
    # additional releases (ar) and their frequency (a_r_i_f),
    # generation when releases start (a_r_g), and how many 
    # generations before each new release (a_r_f). Heatmap
    # adds a few extra details based on how many generations (gens)
    # the data is being averaged over.
    
    from openpyxl import load_workbook
    
    startrow = 9

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startcol is None and sheet_name in writer.book.sheetnames:
            startcol = writer.book[sheet_name].max_column + 1

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startcol is None:
        startcol = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, startcol=startcol, **to_excel_kwargs)
    
    worksheet = writer.sheets[sheet_name]
    mycell = worksheet.cell(row=1, column=startcol+1)
    mycell.value = gene_drive_name
    mycell = worksheet.cell(row=2, column=startcol+1)
    mycell.value=datatype
    mycell = worksheet.cell(row=3, column=startcol+1)
    mycell.value = 'IF = ' + str(i_f)
    mycell = worksheet.cell(row=4, column=startcol+1)
    mycell.value = 'Intro_Geno = ' + str(i_g)
    mycell = worksheet.cell(row=5, column=startcol+1)
    mycell.value = 'cFC = ' + str(f_c_c) + ' per allele'
    mycell = worksheet.cell(row=6, column=startcol+1)
    mycell.value = 'rFC = ' + str(f_c_r) + ' per allele'
    mycell = worksheet.cell(row=7, column=startcol+1)
    mycell.value = str(r_d) + ' cM'
    mycell = worksheet.cell(row=2, column=startcol+3)
    mycell.value = 'Mig_rate = ' + str(r_d)
    if a_r > 0:
        mycell = worksheet.cell(row=3, column=startcol+3)
        mycell.value = str(a_r) + ' ar'
        mycell = worksheet.cell(row=4, column=startcol+3)
        mycell.value = 'arIF = ' +str(a_r_i_f)
        mycell = worksheet.cell(row=5, column=startcol+3)
        mycell.value = ' arGen' + str(a_r_g)
        mycell = worksheet.cell(row=6, column=startcol+3)
        mycell.value = ' arIntro_Geno' + str(a_r_i_g)
        mycell = worksheet.cell(row=7, column=startcol+3)
        mycell.value = ' ar_freq' + str(a_r_f)
    
    if pop_num == 1 or pop_num == 2 or pop_num == 3:
        mycell = worksheet.cell(row=1, column=startcol+3)
        mycell.value = 'Population ' + str(pop_num)
    
    if heatmap:
        mycell = worksheet.cell(row=8, column=startcol+1)
        mycell.value = 'Allele = ' + str(allele)
        mycell = worksheet.cell(row=8, column=startcol+3)
        mycell.value = 'Avg of ' + str(gens) + ' gens'
    
    # save the workbook
    writer.save()


def multi_run(gene_drive_name, master_list, d_a, 
              lethal, rescue, 
              plot_alleles, leg_pos, fig_width, fig_height, 
              num_pops, num_gens, 
              introduction_frequency_list, introduction_type, i_g, 
              introduction_frequency_of_inversion, introduction_type_of_inversion, i_inv_g, 
              f_c_c_list, f_c_r_list, f_c_i_list, f_c_type, 
              r_d_list, r_d_inv_list, 
              m_r_list, 
              a_r_list, additional_release_introduction_frequency_list, additional_release_introduction_type, a_r_i_g, 
              a_r_g_list, a_r_f_list, 
              save=0, file='ClvR data', sheet='default'):
    """
    multi_run takes in the master_list for a given gene drive as well
    as multiple sets of general parameters, then performs, plots, and 
    can save multiple runs of a drive_data" population dynamic 
    simulation.
    
    Parameters as per drive_data, for all parameters with the suffix '_list' (besides 'master_list') just means that there mutliple values to be tested for that parameter. 
    """
    
    # Unpack master_list variables
    num_loci = master_list[0]
    drive_activites_total_counter = master_list[1]
    num_male_geno = master_list[2]
    num_female_geno = master_list[3]
    males_g_a = master_list[4]
    females_g_a = master_list[5]
    male_alleles = master_list[6]
    female_alleles = master_list[7]
    all_alleles = master_list[8]
    male_a_freq_string = master_list[9]
    female_a_freq_string = master_list[10]
    male_p_freq_string = master_list[11]
    female_p_freq_string = master_list[12]
    all_a_freq_string = master_list[13]
    all_p_freq_string = master_list[14]
    d_a_males_string = master_list[15]
    d_a_females_string = master_list[16]
    d_a_males_rates_string = master_list[17]
    d_a_females_rates_string = master_list[18]
    male_cross_matrix = master_list[19]
    female_cross_matrix = master_list[20]
    
    # Preallocate plotting dataframe
    df = pd.DataFrame()
    
    i_f_list = [intro_mod(x, introduction_type) for x in [y*(1-introduction_frequency_of_inversion) for y in introduction_frequency_list]]
    i_f_inv_list = [intro_mod(x, introduction_type_of_inversion) for x in [y*introduction_frequency_of_inversion for y in introduction_frequency_list]]
    a_r_i_f_list = [intro_mod(x, additional_release_introduction_type) for x in additional_release_introduction_frequency_list]
    
    # If r_d doesn't change, evaluate m_eval once
    if len(r_d_list) == 1:
        
        recombination = list(np.zeros(num_loci-1))
        recombination[0] = 1-r_d_list[0]/50
        
        recombination_inv = list(np.zeros(num_loci-1))
        recombination_inv[0] = 1-r_d_inv_list[0]/50
        
        male_cm, female_cm = matrix_evaluator(male_cross_matrix, female_cross_matrix, d_a, 
                                              d_a_males_string, d_a_females_string, 
                                              d_a_males_rates_string, d_a_females_rates_string, 
                                              num_loci, num_male_geno, num_female_geno, 
                                              drive_activites_total_counter, 
                                              recombination, recombination, recombination_inv, recombination_inv)
        
        m_eval = [male_cm, female_cm]
        
    else:
        m_eval = 0
        
    for i_f_ind, i_f in enumerate(i_f_list):
        i_f_inv = i_f_inv_list[i_f_ind]
        for i, f_c_c in enumerate(f_c_c_list):
            f_c_r = f_c_r_list[i]
            f_c_i = f_c_i_list[i]
            for r_d_ind, r_d in enumerate(r_d_list):
                r_d_inv = r_d_inv_list[r_d_ind]
                for m_r in m_r_list:
                    for a_r in a_r_list:
                        for a_r_i_f in a_r_i_f_list:
                            for a_r_g in a_r_g_list:
                                for a_r_f in a_r_f_list:
                                    f_c_male_alleles = np.zeros(len(male_alleles))
                                    f_c_male_alleles[0] = f_c_c
                                    f_c_female_alleles = np.zeros(len(female_alleles))
                                    f_c_female_alleles[0] = f_c_c

                                    if 'I' in all_alleles:
                                        f_c_male_alleles[all_alleles.index('I')] = f_c_i
                                        f_c_female_alleles[all_alleles.index('I')] = f_c_i

                                    if 'R' in all_alleles:
                                        f_c_male_alleles[all_alleles.index('R')] = f_c_r
                                        f_c_female_alleles[all_alleles.index('R')] = f_c_r

                                    f_c = [f_c_male_alleles, f_c_female_alleles]

                                    recombination = list(np.zeros(num_loci-1))
                                    recombination[0] = 1-r_d/50
                                    
                                    recombination_inv = list(np.zeros(num_loci-1))
                                    recombination_inv[0] = 1-r_d_inv/50

                                    run = 'IF' + str(sum(i_f)) + ', cFC'+str(f_c_c) + ', rFC'+str(f_c_r)+', ' + str(r_d) + 'cM'
                                    if a_r > 0:
                                        run += ', ' + str(a_r) + 'ar, ' + str(a_r_i_f) + 'arIF '+str(a_r_g) + 'arg, ' + str(a_r_f) + 'arf'

                                    if num_pops == 1:

                                        all_p_f_1, all_a_f_1 = drive_data_1pop(master_list, d_a, lethal, rescue, 
                                                                               m_eval, num_gens, 
                                                                               i_f, i_g, 
                                                                               i_f_inv, i_inv_g, 
                                                                               f_c, f_c_type, 
                                                                               recombination, recombination, recombination_inv, recombination_inv, 
                                                                               a_r, a_r_i_f, a_r_i_g, 
                                                                               a_r_g, a_r_f)

                                        df = add_new_data_to_dataframe(df, all_p_f_1, run, 'phenotype', 1, num_gens, all_alleles)
                                        df = add_new_data_to_dataframe(df, all_a_f_1, run, 'allele', 1, num_gens, all_alleles)

                                        # If saving, add each set of data as its generated
                                        if save == 1:
                                            df_p = pd.DataFrame(all_p_f_1, columns = all_alleles)
                                            df_a = pd.DataFrame(all_a_f_1, columns = all_alleles)

                                            append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                                               df_p, 0, 
                                                               gene_drive_name, '\'Phenotype\' frequency', 
                                                               i_f, i_g,
                                                               f_c_c, f_c_r, 
                                                               r_d, 
                                                               0.0, 
                                                               a_r, a_r_i_f, a_r_i_g, 
                                                               a_r_g, a_r_f)
                                            append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                                               df_a, 0, 
                                                               gene_drive_name, 'Allele frequency', 
                                                               i_f, i_g,
                                                               f_c_c, f_c_r, 
                                                               r_d, 
                                                               0.0, 
                                                               a_r, a_r_i_f, a_r_i_g, 
                                                               a_r_g, a_r_f)

                                    elif num_pops == 2:

                                        all_p_f_1, all_a_f_1, \
                                        all_p_f_2, all_a_f_2 = drive_data_2pop(master_list, d_a, lethal, rescue, 
                                                                               m_eval, num_gens, 
                                                                               i_f, i_g, 
                                                                               i_f_inv, i_inv_g, 
                                                                               f_c, f_c_type, 
                                                                               recombination, recombination, recombination_inv, recombination_inv, 
                                                                               m_r, 
                                                                               a_r, a_r_i_f, a_r_i_g, 
                                                                               a_r_g, a_r_f)

                                        df = add_new_data_to_dataframe(df, all_p_f_1, run, 'phenotype', 1, num_gens, all_alleles)
                                        df = add_new_data_to_dataframe(df, all_a_f_1, run, 'allele', 1, num_gens, all_alleles)

                                        df = add_new_data_to_dataframe(df, all_p_f_2, run, 'phenotype', 2, num_gens, all_alleles)
                                        df = add_new_data_to_dataframe(df, all_a_f_2, run, 'allele', 2, num_gens, all_alleles)

                                        # If saving, add each set of data as its generated
                                        if save == 1:
                                            df_p = pd.DataFrame(all_p_f_1, columns = all_alleles)
                                            df_a = pd.DataFrame(all_a_f_1, columns = all_alleles)

                                            append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                                               df_p, 1, 
                                                               gene_drive_name, '\'Phenotype\' frequency', 
                                                               i_f, i_g,
                                                               f_c_c, f_c_r, 
                                                               r_d, 
                                                               m_r, 
                                                               a_r, a_r_i_f, a_r_i_g, 
                                                               a_r_g, a_r_f)
                                            append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                                               df_a, 1, 
                                                               gene_drive_name, 'Allele frequency', 
                                                               i_f, i_g,
                                                               f_c_c, f_c_r, 
                                                               r_d, 
                                                               m_r, 
                                                               a_r, a_r_i_f, a_r_i_g, 
                                                               a_r_g, a_r_f)

                                            df_p = pd.DataFrame(all_p_f_2, columns = all_alleles)
                                            df_a = pd.DataFrame(all_a_f_2, columns = all_alleles)

                                            append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                                               df_p, 2, 
                                                               gene_drive_name, '\'Phenotype\' frequency', 
                                                               i_f, i_g,
                                                               f_c_c, f_c_r, 
                                                               r_d, 
                                                               m_r, 
                                                               a_r, a_r_i_f, a_r_i_g, 
                                                               a_r_g, a_r_f)
                                            append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                                               df_a, 2, 
                                                               gene_drive_name, 'Allele frequency', 
                                                               i_f, i_g,
                                                               f_c_c, f_c_r, 
                                                               r_d, 
                                                               m_r, 
                                                               a_r, a_r_i_f, a_r_i_g, 
                                                               a_r_g, a_r_f)

                                    elif num_pops == 3:

                                        all_p_f_1, all_a_f_1, \
                                        all_p_f_2, all_a_f_2, \
                                        all_p_f_3, all_a_f_3 = drive_data_3pop(master_list, d_a, lethal, rescue, 
                                                                               m_eval, num_gens, 
                                                                               i_f, i_g, 
                                                                               i_f_inv, i_inv_g, 
                                                                               f_c, f_c_type, 
                                                                               recombination, recombination, recombination_inv, recombination_inv, 
                                                                               m_r, 
                                                                               a_r, a_r_i_f, a_r_i_g,
                                                                               a_r_g, a_r_f)

                                        df = add_new_data_to_dataframe(df, all_p_f_1, run, 'phenotype', 1, num_gens, all_alleles)
                                        df = add_new_data_to_dataframe(df, all_a_f_1, run, 'allele', 1, num_gens, all_alleles)

                                        df = add_new_data_to_dataframe(df, all_p_f_2, run, 'phenotype', 2, num_gens, all_alleles)
                                        df = add_new_data_to_dataframe(df, all_a_f_2, run, 'allele', 2, num_gens, all_alleles)

                                        df = add_new_data_to_dataframe(df, all_p_f_3, run, 'phenotype', 3, num_gens, all_alleles)
                                        df = add_new_data_to_dataframe(df, all_a_f_3, run, 'allele', 3, num_gens, all_alleles)

                                        # If saving, add each set of data as its generated
                                        if save == 1:
                                            df_p = pd.DataFrame(all_p_f_1, columns = all_alleles)
                                            df_a = pd.DataFrame(all_a_f_1, columns = all_alleles)

                                            append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                                               df_p, 1, 
                                                               gene_drive_name, '\'Phenotype\' frequency', 
                                                               i_f, i_g,
                                                               f_c_c, f_c_r, 
                                                               r_d, 
                                                               m_r, 
                                                               a_r, a_r_i_f, a_r_i_g, 
                                                               a_r_g, a_r_f)
                                            append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                                               df_a, 1, 
                                                               gene_drive_name, 'Allele frequency', 
                                                               i_f, i_g,
                                                               f_c_c, f_c_r, 
                                                               r_d, 
                                                               m_r, 
                                                               a_r, a_r_i_f, a_r_i_g, 
                                                               a_r_g, a_r_f)

                                            df_p = pd.DataFrame(all_p_f_2, columns = all_alleles)
                                            df_a = pd.DataFrame(all_a_f_2, columns = all_alleles)

                                            append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                                               df_p, 2, 
                                                               gene_drive_name, '\'Phenotype\' frequency', 
                                                               i_f, i_g,
                                                               f_c_c, f_c_r, 
                                                               r_d, 
                                                               m_r, 
                                                               a_r, a_r_i_f, a_r_i_g, 
                                                               a_r_g, a_r_f)
                                            append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                                               df_a, 2, 
                                                               gene_drive_name, 'Allele frequency', 
                                                               i_f, i_g,
                                                               f_c_c, f_c_r, 
                                                               r_d, 
                                                               m_r, 
                                                               a_r, a_r_i_f, a_r_i_g, 
                                                               a_r_g, a_r_f)

                                            df_p = pd.DataFrame(all_p_f_3, columns = all_alleles)
                                            df_a = pd.DataFrame(all_a_f_3, columns = all_alleles)

                                            append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                                               df_p, 3, 
                                                               gene_drive_name, '\'Phenotype\' frequency', 
                                                               i_f, i_g,
                                                               f_c_c, f_c_r, 
                                                               r_d, 
                                                               m_r, 
                                                               a_r, a_r_i_f, a_r_i_g, 
                                                               a_r_g, a_r_f)
                                            append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                                               df_a, 3, 
                                                               gene_drive_name, 'Allele frequency', 
                                                               i_f, i_g,
                                                               f_c_c, f_c_r, 
                                                               r_d, 
                                                               m_r, 
                                                               a_r, a_r_i_f, a_r_i_g, 
                                                               a_r_g, a_r_f)

                                    else:
                                        raise RuntimeError(f'num_pops must be 1, 2, or 3.')
    
    temp_plot = multi_run_plots(df, gene_drive_name, num_pops, plot_alleles, leg_pos, fig_width, fig_height)
    
    if num_pops == 1:
        data_plot = temp_plot.cols(1)

    elif num_pops == 2:
        data_plot = temp_plot.cols(2)

    else:
        data_plot = temp_plot.cols(3)
        
    return(df, data_plot)


def intro_mod(i_f, intro_type='both'):
    """
    intro_mod modifies introduction frequency from a single
    value into a pair based on which genders are being
    introduced (via intro_type).
    """
    
    if intro_type == 'both':
        return([i_f/2, i_f/2])
    
    elif intro_type == 'males':
        return([i_f, 0.0])
    
    elif intro_type == 'females':
        return([0.0, i_f])
    
    else:
        RuntimeError(f'intro_type must be \'males\', \'females\', or \'both\'')


def add_new_data_to_dataframe(df, data, run, freq_type, pop, num_gens, all_alleles):
    """
    add_new_data_to_dataframe takes new data and combines it with the 
    previous data into a single dataframe.
    """
    
    df_temp = pd.DataFrame(data, columns = all_alleles)
    df_temp = df_temp.melt()
    df_temp['Generation'] = [int(gen) for gen in np.linspace(0, num_gens, num_gens+1)]*len(all_alleles)
    df_temp = df_temp.rename(columns={"variable": "Allele", "value": "Frequency"})
    df_temp['Frequency Type'] = freq_type
    df_temp['Run'] = run
    df_temp['Population'] = pop
    
    df = df.append(df_temp)
    
    return(df)

def multi_run_plots(df, gene_drive_name, num_pops, plot_alleles, leg_pos, fig_width, fig_height):
    """
    multi_run_plots handles plotting for the multi_run function.
    """
    
    plot_list = []
    
    for allele in plot_alleles:
        for pop in range(1, num_pops+1):
            
            plot = hv.Points(
                data=df[(df['Allele'] == allele) & (df['Frequency Type'] == 'phenotype') & (df['Population'] == pop)],
                kdims=['Generation', 'Frequency'],
                vdims=['Run'],
            ).groupby(
                'Run'
            ).opts(
                title=str(gene_drive_name) + ' Frequency of ' + str(allele) + ' in population ' + str(pop),
                tools=['hover'],
                legend_position=leg_pos,
                width=fig_width,
                height=fig_height
            ).overlay(
            )

            plot_list.append(plot)
        
    data_plot = plot_list[0]

    if len(plot_list) >= 2:
        for p in range(1, len(plot_list)):
            data_plot += plot_list[p]

    return(data_plot)


def heatmap(gene_drive_name, h_type, 
            master_list, d_a, 
            lethal, rescue, 
            num_pops, num_gens, num_gen_avg, plot_alleles, 
            figure_width, figure_height, 
            introduction_frequency_list, introduction_type, i_g, 
            introduction_frequency_of_inversion, introduction_type_of_inversion, i_inv_g,
            f_c_c_list, f_c_r_list, f_c_i_list, f_c_type, 
            r_d_list, r_d_inv_list, 
            m_r_list, 
            a_r, additional_release_introduction_frequency, 
            additional_release_introduction_type, a_r_i_g, 
            a_r_g, a_r_f,
            file='ClvR data', sheet='default'):
    """
    Like multi_run, performs multiple single runs but does so over a 
    range of values for two parameters and holding the rest fixed in 
    order to generate a heatmap (color indicates average frequency 
    over a multiple of 100 generations).
    """
    
    # Unpack master_list variables
    num_loci = master_list[0]
    drive_activites_total_counter = master_list[1]
    num_male_geno = master_list[2]
    num_female_geno = master_list[3]
    males_g_a = master_list[4]
    females_g_a = master_list[5]
    male_alleles = master_list[6]
    female_alleles = master_list[7]
    all_alleles = master_list[8]
    male_a_freq_string = master_list[9]
    female_a_freq_string = master_list[10]
    male_p_freq_string = master_list[11]
    female_p_freq_string = master_list[12]
    all_a_freq_string = master_list[13]
    all_p_freq_string = master_list[14]
    d_a_males_string = master_list[15]
    d_a_females_string = master_list[16]
    d_a_males_rates_string = master_list[17]
    d_a_females_rates_string = master_list[18]
    male_cross_matrix = master_list[19]
    female_cross_matrix = master_list[20]
    
    df = pd.DataFrame()
    
    i_f_list = [intro_mod(x, introduction_type) for x in [y*(1-introduction_frequency_of_inversion) for y in introduction_frequency_list]]
    i_f_inv_list = [intro_mod(x, introduction_type_of_inversion) for x in [y*introduction_frequency_of_inversion for y in introduction_frequency_list]]
    a_r_i_f = intro_mod(additional_release_introduction_frequency, additional_release_introduction_type)
    
    i_f_steps = len(i_f_list)
    f_c_steps = len(f_c_c_list)
    r_d_steps = len(r_d_list)
    m_r_steps = len(m_r_list)
        
    gen_level = int(np.ceil(num_gens/100))
    
    if num_pops == 1:
        
        if h_type == 'FCvIF':
            
            phenotypes_matrix = np.array([[[[None]*(f_c_steps)]*(i_f_steps)]*len(all_alleles)]*gen_level)
            alleles_matrix = np.array([[[[None]*(f_c_steps)]*(i_f_steps)]*len(all_alleles)]*gen_level)
            
            cols = ['Fitness Cost', 
                    'Introduction Frequency', 
                    'Averaged Frequency', 
                    'Frequency Type', 
                    'Number of Generations Data Averaged Over', 
                    'Allele', 
                    'Population']
            
            recombination = list(np.zeros(num_loci-1))
            recombination[0] = 1-r_d_list[0]/50
            
            recombination_inv = list(np.zeros(num_loci-1))
            recombination_inv[0] = 1-r_d_inv_list[0]/50

            male_cm, female_cm = matrix_evaluator(male_cross_matrix, female_cross_matrix, d_a, 
                                                  d_a_males_string, d_a_females_string, 
                                                  d_a_males_rates_string, d_a_females_rates_string, 
                                                  num_loci, num_male_geno, num_female_geno, 
                                                  drive_activites_total_counter, 
                                                  recombination, recombination, recombination_inv, recombination_inv)

            m_eval = [male_cm, female_cm]
            
            for i_f_i, i_f in enumerate(i_f_list):
                i_f_inv = i_f_inv_list[i_f_i]
                for f_c_ind, f_c_c in enumerate(f_c_c_list):
                    f_c_r = f_c_r_list[f_c_ind]
                    f_c_i = f_c_r_list[f_c_ind]
                    f_c_male_alleles = np.zeros(len(male_alleles))
                    f_c_male_alleles[0] = f_c_c
                    f_c_female_alleles = np.zeros(len(female_alleles))
                    f_c_female_alleles[0] = f_c_c

                    if 'I' in all_alleles:
                        f_c_male_alleles[all_alleles.index('I')] = f_c_i
                        f_c_female_alleles[all_alleles.index('I')] = f_c_i

                    if 'R' in all_alleles:
                        f_c_male_alleles[all_alleles.index('R')] = f_c_r
                        f_c_female_alleles[all_alleles.index('R')] = f_c_r

                    f_c = [f_c_male_alleles, f_c_female_alleles]
                    
                    all_p_f_1, all_a_f_1 = drive_data_1pop(master_list, d_a, lethal, rescue, 
                                                           m_eval, num_gens, 
                                                           i_f, i_g, 
                                                           i_f_inv, i_inv_g, 
                                                           f_c, f_c_type, 
                                                           recombination, recombination, recombination_inv, recombination_inv, 
                                                           a_r, a_r_i_f, a_r_i_g, 
                                                           a_r_g, a_r_f)
                    
                    for level in range(1, gen_level+1):
                        p_f_ave_1 = np.mean(all_p_f_1[:level*100], axis=0)
                        a_f_ave_1 = np.mean(all_a_f_1[:level*100], axis=0)

                        for a_i, a in enumerate(all_alleles):
                            data = [[f_c_c, np.sum(i_f), p_f_ave_1[a_i], 'phenotype', level*100, a, 1],
                                    [f_c_c, np.sum(i_f), a_f_ave_1[a_i], 'allele', level*100, a, 1]]
                            
                            df = df.append(pd.DataFrame(data, columns = cols))
                            
                            phenotypes_matrix[level-1, a_i, i_f_i, f_c_ind] = p_f_ave_1[a_i]
                            alleles_matrix[level-1, a_i, i_f_i, f_c_ind] = a_f_ave_1[a_i]

        elif h_type == 'RDvIF':
            
            phenotypes_matrix = np.array([[[[None]*(r_d_steps)]*(i_f_steps)]*len(all_alleles)]*gen_level)
            alleles_matrix = np.array([[[[None]*(r_d_steps)]*(i_f_steps)]*len(all_alleles)]*gen_level)
            
            cols = ['Recombination Distance', 
                    'Introduction Frequency', 
                    'Averaged Frequency', 
                    'Frequency Type', 
                    'Number of Generations Data Averaged Over', 
                    'Allele', 
                    'Population']
            
            f_c_c = f_c_c_list[0]
            f_c_r = f_c_r_list[0]
            f_c_i = f_c_i_list[0]
            
            f_c_male_alleles = np.zeros(len(male_alleles))
            f_c_male_alleles[0] = f_c_c
            f_c_female_alleles = np.zeros(len(female_alleles))
            f_c_female_alleles[0] = f_c_c

            if 'I' in all_alleles:
                f_c_male_alleles[all_alleles.index('I')] = f_c_i
                f_c_female_alleles[all_alleles.index('I')] = f_c_i

            if 'R' in all_alleles:
                f_c_male_alleles[all_alleles.index('R')] = f_c_r
                f_c_female_alleles[all_alleles.index('R')] = f_c_r

            f_c = [f_c_male_alleles, f_c_female_alleles]
            
            for i_f_i, i_f in enumerate(i_f_list):
                i_f_inv = i_f_inv_list[i_f_i]
                for r_d_i, r_d in enumerate(r_d_list):
                    
                    recombination = list(np.zeros(num_loci-1))
                    recombination[0] = 1-r_d/50

                    recombination_inv = list(np.zeros(num_loci-1))
                    recombination_inv[0] = 1-r_d_inv_list[r_d_i]/50
                    
                    all_p_f_1, all_a_f_1 = drive_data_1pop(master_list, d_a, lethal, rescue, 
                                                           0, num_gens, 
                                                           i_f, i_g, 
                                                           i_f_inv, i_inv_g, 
                                                           f_c, f_c_type, 
                                                           recombination, recombination, recombination_inv, recombination_inv, 
                                                           a_r, a_r_i_f, a_r_i_g, 
                                                           a_r_g, a_r_f)
                    for level in range(1, gen_level+1):
                        p_f_ave_1 = np.mean(all_p_f_1[:level*100], axis=0)
                        a_f_ave_1 = np.mean(all_a_f_1[:level*100], axis=0)

                        for a_i, a in enumerate(all_alleles):
                            data = [[r_d, np.sum(i_f), p_f_ave_1[a_i], 'phenotype', level*100, a, 1],
                                    [r_d, np.sum(i_f), a_f_ave_1[a_i], 'allele', level*100, a, 1]]
                            
                            df = df.append(pd.DataFrame(data, columns = cols))
                            
                            phenotypes_matrix[level-1, a_i, i_f_i, r_d_i] = p_f_ave_1[a_i]
                            alleles_matrix[level-1, a_i, i_f_i, r_d_i] = a_f_ave_1[a_i]

        elif h_type == 'FCvRD':
            
            phenotypes_matrix = np.array([[[[None]*(f_c_steps)]*(r_d_steps)]*len(all_alleles)]*gen_level)
            alleles_matrix = np.array([[[[None]*(f_c_steps)]*(r_d_steps)]*len(all_alleles)]*gen_level)
            
            cols = ['Fitness Cost', 
                    'Recombination Distance', 
                    'Averaged Frequency', 
                    'Frequency Type', 
                    'Number of Generations Data Averaged Over', 
                    'Allele', 
                    'Population']
            
            i_f = i_f_list[0]
            i_f_inv = i_f_inv_list[0]
            
            for r_d_i, r_d in enumerate(r_d_list):
                for f_c_ind, f_c_c in enumerate(f_c_c_list):
                    f_c_r = f_c_r_list[f_c_ind]
                    f_c_i = f_c_r_list[f_c_ind]
                    f_c_male_alleles = np.zeros(len(male_alleles))
                    f_c_male_alleles[0] = f_c_c
                    f_c_female_alleles = np.zeros(len(female_alleles))
                    f_c_female_alleles[0] = f_c_c

                    if 'I' in all_alleles:
                        f_c_male_alleles[all_alleles.index('I')] = f_c_i
                        f_c_female_alleles[all_alleles.index('I')] = f_c_i

                    if 'R' in all_alleles:
                        f_c_male_alleles[all_alleles.index('R')] = f_c_r
                        f_c_female_alleles[all_alleles.index('R')] = f_c_r

                    f_c = [f_c_male_alleles, f_c_female_alleles]
                    
                    recombination = list(np.zeros(num_loci-1))
                    recombination[0] = 1-r_d/50

                    recombination_inv = list(np.zeros(num_loci-1))
                    recombination_inv[0] = 1-r_d_inv_list[r_d_i]/50
                    
                    all_p_f_1, all_a_f_1 = drive_data_1pop(master_list, d_a, lethal, rescue, 
                                                           0, num_gens, 
                                                           i_f, i_g, 
                                                           i_f_inv, i_inv_g, 
                                                           f_c, f_c_type, 
                                                           recombination, recombination, recombination_inv, recombination_inv, 
                                                           a_r, a_r_i_f, a_r_i_g, 
                                                           a_r_g, a_r_f)
                    for level in range(1, gen_level+1):
                        p_f_ave_1 = np.mean(all_p_f_1[:level*100], axis=0)
                        a_f_ave_1 = np.mean(all_a_f_1[:level*100], axis=0)

                        for a_i, a in enumerate(all_alleles):
                            data = [[f_c_c, r_d, p_f_ave_1[a_i], 'phenotype', level*100, a, 1],
                                    [f_c_c, r_d, a_f_ave_1[a_i], 'allele', level*100, a, 1]]
                            
                            df = df.append(pd.DataFrame(data, columns = cols))
                            
                            phenotypes_matrix[level-1, a_i, r_d_i, f_c_ind] = p_f_ave_1[a_i]
                            alleles_matrix[level-1, a_i, r_d_i, f_c_ind] = a_f_ave_1[a_i]
        
        sheet_name = sheet + '_1pop'

        for level in range(1, gen_level+1):
            for a_i, a in enumerate(all_alleles):
                df_p_1 = pd.DataFrame(phenotypes_matrix[level-1, a_i])
                df_a_1 = pd.DataFrame(alleles_matrix[level-1, a_i])
                
                append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                   df_p_1, 0, 
                                   gene_drive_name, '\'Phenotype\' frequency', 
                                   i_f_list, i_g, 
                                   f_c_c_list, f_c_r_list, 
                                   r_d_list, 
                                   0.0, 
                                   a_r, a_r_i_f, a_r_i_g, 
                                   a_r_g, a_r_f, 
                                   1, a, level*100)
                append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                   df_a_1, 0, 
                                   gene_drive_name, 'Allele frequency',
                                   i_f_list, i_g, 
                                   f_c_c_list, f_c_r_list, 
                                   r_d_list, 
                                   0.0, 
                                   a_r, a_r_i_f, a_r_i_g, 
                                   a_r_g, a_r_f, 
                                   1, a, level*100)
        
        # Plot Heatmap(s)
        temp_plot = heatmap_plots(gene_drive_name, h_type, df, num_pops, plot_alleles,  num_gen_avg, figure_width, figure_height)

        if num_pops == 1:
            data_plot = temp_plot.cols(1)

        elif num_pops == 2:
            data_plot = temp_plot.cols(2)

        else:
            data_plot = temp_plot.cols(3)
        
        return(df, data_plot)
        
    elif num_pops == 2:
        
        if h_type == 'FCvIF':
            
            phenotypes_matrix = np.array([[[[[None]*(f_c_steps)]*(i_f_steps)]*len(all_alleles)]*gen_level]*num_pops)
            alleles_matrix = np.array([[[[[None]*(f_c_steps)]*(i_f_steps)]*len(all_alleles)]*gen_level]*num_pops)
            
            cols = ['Fitness Cost', 
                    'Introduction Frequency', 
                    'Averaged Frequency', 
                    'Frequency Type', 
                    'Number of Generations Data Averaged Over', 
                    'Allele', 
                    'Population']
            
            recombination = list(np.zeros(num_loci-1))
            recombination[0] = 1-r_d_list[0]/50

            recombination_inv = list(np.zeros(num_loci-1))
            recombination_inv[0] = 1-r_d_inv_list[0]/50

            male_cm, female_cm = matrix_evaluator(male_cross_matrix, female_cross_matrix, d_a, 
                                                  d_a_males_string, d_a_females_string, 
                                                  d_a_males_rates_string, d_a_females_rates_string, 
                                                  num_loci, num_male_geno, num_female_geno, 
                                                  drive_activites_total_counter, 
                                                  recombination, recombination, recombination_inv, recombination_inv)

            m_eval = [male_cm, female_cm]
            
            m_r = m_r_list[0]
            
            for i_f_i, i_f in enumerate(i_f_list):
                i_f_inv = i_f_inv_list[i_f_i]
                for f_c_ind, f_c_c in enumerate(f_c_c_list):
                    f_c_r = f_c_r_list[f_c_ind]
                    f_c_i = f_c_r_list[f_c_ind]
                    f_c_male_alleles = np.zeros(len(male_alleles))
                    f_c_male_alleles[0] = f_c_c
                    f_c_female_alleles = np.zeros(len(female_alleles))
                    f_c_female_alleles[0] = f_c_c

                    if 'I' in all_alleles:
                        f_c_male_alleles[all_alleles.index('I')] = f_c_i
                        f_c_female_alleles[all_alleles.index('I')] = f_c_i

                    if 'R' in all_alleles:
                        f_c_male_alleles[all_alleles.index('R')] = f_c_r
                        f_c_female_alleles[all_alleles.index('R')] = f_c_r

                    f_c = [f_c_male_alleles, f_c_female_alleles]
                    
                    all_p_f_1, all_a_f_1, \
                    all_p_f_2, all_a_f_2 = drive_data_2pop(master_list, d_a, lethal, rescue, 
                                                           m_eval, num_gens, 
                                                           i_f, i_g, 
                                                           i_f_inv, i_inv_g, 
                                                           f_c, f_c_type, 
                                                           recombination, recombination, recombination_inv, recombination_inv, 
                                                           m_r, 
                                                           a_r, a_r_i_f, a_r_i_g, 
                                                           a_r_g, a_r_f)
                    for level in range(1, gen_level+1):
                        p_f_ave_1 = np.mean(all_p_f_1[:level*100], axis=0)
                        a_f_ave_1 = np.mean(all_a_f_1[:level*100], axis=0)
                        
                        p_f_ave_2 = np.mean(all_p_f_2[:level*100], axis=0)
                        a_f_ave_2 = np.mean(all_a_f_2[:level*100], axis=0)

                        for a_i, a in enumerate(all_alleles):
                            data = [[f_c_c, np.sum(i_f), p_f_ave_1[a_i], 'phenotype', level*100, a, 1],
                                    [f_c_c, np.sum(i_f), a_f_ave_1[a_i], 'allele', level*100, a, 1], 
                                    [f_c_c, np.sum(i_f), p_f_ave_2[a_i], 'phenotype', level*100, a, 2],
                                    [f_c_c, np.sum(i_f), a_f_ave_2[a_i], 'allele', level*100, a, 2]]
                            
                            df = df.append(pd.DataFrame(data, columns = cols))
                            
                            phenotypes_matrix[0, level-1, a_i, i_f_i, f_c_ind] = p_f_ave_1[a_i]
                            alleles_matrix[0, level-1, a_i, i_f_i, f_c_ind] = a_f_ave_1[a_i]
                            
                            phenotypes_matrix[1, level-1, a_i, i_f_i, f_c_ind] = p_f_ave_2[a_i]
                            alleles_matrix[1, level-1, a_i, i_f_i, f_c_ind] = a_f_ave_2[a_i]

        elif h_type == 'RDvIF':
            
            phenotypes_matrix = np.array([[[[[None]*(r_d_steps)]*(i_f_steps)]*len(all_alleles)]*gen_level]*num_pops)
            alleles_matrix = np.array([[[[[None]*(r_d_steps)]*(i_f_steps)]*len(all_alleles)]*gen_level]*num_pops)
            
            cols = ['Recombination Distance', 
                    'Introduction Frequency', 
                    'Averaged Frequency', 
                    'Frequency Type', 
                    'Number of Generations Data Averaged Over', 
                    'Allele', 
                    'Population']
            
            f_c_c = f_c_c_list[0]
            f_c_r = f_c_r_list[0]
            f_c_i = f_c_i_list[0]
            
            f_c_male_alleles = np.zeros(len(male_alleles))
            f_c_male_alleles[0] = f_c_c
            f_c_female_alleles = np.zeros(len(female_alleles))
            f_c_female_alleles[0] = f_c_c

            if 'I' in all_alleles:
                f_c_male_alleles[all_alleles.index('I')] = f_c_i
                f_c_female_alleles[all_alleles.index('I')] = f_c_i

            if 'R' in all_alleles:
                f_c_male_alleles[all_alleles.index('R')] = f_c_r
                f_c_female_alleles[all_alleles.index('R')] = f_c_r

            f_c = [f_c_male_alleles, f_c_female_alleles]
            
            m_r = m_r_list[0]
            
            for i_f_i, i_f in enumerate(i_f_list):
                i_f_inv = i_f_inv_list[i_f_i]
                for r_d_i, r_d in enumerate(r_d_list):
                    
                    recombination = list(np.zeros(num_loci-1))
                    recombination[0] = 1-r_d/50

                    recombination_inv = list(np.zeros(num_loci-1))
                    recombination_inv[0] = 1-r_d_inv_list[r_d_i]/50
                    
                    all_p_f_1, all_a_f_1, \
                    all_p_f_2, all_a_f_2 = drive_data_2pop(master_list, d_a, lethal, rescue, 
                                                           0, num_gens, 
                                                           i_f, i_g, 
                                                           i_f_inv, i_inv_g, 
                                                           f_c, f_c_type, 
                                                           recombination, recombination, recombination_inv, recombination_inv, 
                                                           m_r, 
                                                           a_r, a_r_i_f, a_r_i_g, 
                                                           a_r_g, a_r_f)
                    for level in range(1, gen_level+1):
                        p_f_ave_1 = np.mean(all_p_f_1[:level*100], axis=0)
                        a_f_ave_1 = np.mean(all_a_f_1[:level*100], axis=0)
                        
                        p_f_ave_2 = np.mean(all_p_f_2[:level*100], axis=0)
                        a_f_ave_2 = np.mean(all_a_f_2[:level*100], axis=0)

                        for a_i, a in enumerate(all_alleles):
                            data = [[r_d, np.sum(i_f), p_f_ave_1[a_i], 'phenotype', level*100, a, 1],
                                    [r_d, np.sum(i_f), a_f_ave_1[a_i], 'allele', level*100, a, 1], 
                                    [r_d, np.sum(i_f), p_f_ave_2[a_i], 'phenotype', level*100, a, 2],
                                    [r_d, np.sum(i_f), a_f_ave_2[a_i], 'allele', level*100, a, 2]]
                            
                            df = df.append(pd.DataFrame(data, columns = cols))
                            
                            phenotypes_matrix[0, level-1, a_i, i_f_i, r_d_i] = p_f_ave_1[a_i]
                            alleles_matrix[0, level-1, a_i, i_f_i, r_d_i] = a_f_ave_1[a_i]
                            
                            phenotypes_matrix[1, level-1, a_i, i_f_i, r_d_i] = p_f_ave_2[a_i]
                            alleles_matrix[1, level-1, a_i, i_f_i, r_d_i] = a_f_ave_2[a_i]

        elif h_type == 'FCvRD':
            
            phenotypes_matrix = np.array([[[[[None]*(f_c_steps)]*(r_d_steps)]*len(all_alleles)]*gen_level]*num_pops)
            alleles_matrix = np.array([[[[[None]*(f_c_steps)]*(r_d_steps)]*len(all_alleles)]*gen_level]*num_pops)
            
            cols = ['Fitness Cost', 
                    'Recombination Distance', 
                    'Averaged Frequency', 
                    'Frequency Type', 
                    'Number of Generations Data Averaged Over', 
                    'Allele', 
                    'Population']
            
            i_f = i_f_list[0]
            i_f_inv = i_f_inv_list[0]
            
            m_r = m_r_list[0]
            
            for r_d_i, r_d in enumerate(r_d_list):
                for f_c_ind, f_c_c in enumerate(f_c_c_list):
                    f_c_r = f_c_r_list[f_c_ind]
                    f_c_i = f_c_r_list[f_c_ind]
                    f_c_male_alleles = np.zeros(len(male_alleles))
                    f_c_male_alleles[0] = f_c_c
                    f_c_female_alleles = np.zeros(len(female_alleles))
                    f_c_female_alleles[0] = f_c_c

                    if 'I' in all_alleles:
                        f_c_male_alleles[all_alleles.index('I')] = f_c_i
                        f_c_female_alleles[all_alleles.index('I')] = f_c_i

                    if 'R' in all_alleles:
                        f_c_male_alleles[all_alleles.index('R')] = f_c_r
                        f_c_female_alleles[all_alleles.index('R')] = f_c_r

                    f_c = [f_c_male_alleles, f_c_female_alleles]
                    
                    recombination = list(np.zeros(num_loci-1))
                    recombination[0] = 1-r_d/50

                    recombination_inv = list(np.zeros(num_loci-1))
                    recombination_inv[0] = 1-r_d_inv_list[r_d_i]/50
                    
                    all_p_f_1, all_a_f_1, \
                    all_p_f_2, all_a_f_2 = drive_data_2pop(master_list, d_a, lethal, rescue, 
                                                           0, num_gens, 
                                                           i_f, i_g, 
                                                           i_f_inv, i_inv_g, 
                                                           f_c, f_c_type, 
                                                           recombination, recombination, recombination_inv, recombination_inv, 
                                                           m_r, 
                                                           a_r, a_r_i_f, a_r_i_g, 
                                                           a_r_g, a_r_f)
                    for level in range(1, gen_level+1):
                        p_f_ave_1 = np.mean(all_p_f_1[:level*100], axis=0)
                        a_f_ave_1 = np.mean(all_a_f_1[:level*100], axis=0)
                        
                        p_f_ave_2 = np.mean(all_p_f_2[:level*100], axis=0)
                        a_f_ave_2 = np.mean(all_a_f_2[:level*100], axis=0)

                        for a_i, a in enumerate(all_alleles):
                            data = [[f_c_c, r_d, p_f_ave_1[a_i], 'phenotype', level*100, a, 1],
                                    [f_c_c, r_d, a_f_ave_1[a_i], 'allele', level*100, a, 1], 
                                    [f_c_c, r_d, p_f_ave_2[a_i], 'phenotype', level*100, a, 2],
                                    [f_c_c, r_d, a_f_ave_2[a_i], 'allele', level*100, a, 2]]
                            
                            df = df.append(pd.DataFrame(data, columns = cols))
                            
                            phenotypes_matrix[0, level-1, a_i, r_d_i, f_c_ind] = p_f_ave_1[a_i]
                            alleles_matrix[0, level-1, a_i, r_d_i, f_c_ind] = a_f_ave_1[a_i]
                            
                            phenotypes_matrix[1, level-1, a_i, r_d_i, f_c_ind] = p_f_ave_2[a_i]
                            alleles_matrix[1, level-1, a_i, r_d_i, f_c_ind] = a_f_ave_2[a_i]

        elif h_type == 'MRvRD':
            
            phenotypes_matrix = np.array([[[[[None]*(m_r_steps)]*(r_d_steps)]*len(all_alleles)]*gen_level]*num_pops)
            alleles_matrix = np.array([[[[[None]*(m_r_steps)]*(r_d_steps)]*len(all_alleles)]*gen_level]*num_pops)
            
            cols = ['Migration Rate', 
                    'Recombination Distance', 
                    'Averaged Frequency', 
                    'Frequency Type', 
                    'Number of Generations Data Averaged Over', 
                    'Allele', 
                    'Population']
            
            i_f = i_f_list[0]
            i_f_inv = i_f_inv_list[0]
            
            f_c_c = f_c_c_list[0]
            f_c_r = f_c_r_list[0]
            f_c_i = f_c_i_list[0]
            
            f_c_male_alleles = np.zeros(len(male_alleles))
            f_c_male_alleles[0] = f_c_c
            f_c_female_alleles = np.zeros(len(female_alleles))
            f_c_female_alleles[0] = f_c_c

            if 'I' in all_alleles:
                f_c_male_alleles[all_alleles.index('I')] = f_c_i
                f_c_female_alleles[all_alleles.index('I')] = f_c_i

            if 'R' in all_alleles:
                f_c_male_alleles[all_alleles.index('R')] = f_c_r
                f_c_female_alleles[all_alleles.index('R')] = f_c_r

            f_c = [f_c_male_alleles, f_c_female_alleles]
            
            for r_d_i, r_d in enumerate(r_d_list):
                for m_r_i, m_r in enumerate(m_r_list):
                    
                    recombination = list(np.zeros(num_loci-1))
                    recombination[0] = 1-r_d/50

                    recombination_inv = list(np.zeros(num_loci-1))
                    recombination_inv[0] = 1-r_d_inv_list[r_d_i]/50
                    
                    all_p_f_1, all_a_f_1, \
                    all_p_f_2, all_a_f_2 = drive_data_2pop(master_list, d_a, lethal, rescue, 
                                                           0, num_gens, 
                                                           i_f, i_g, 
                                                           i_f_inv, i_inv_g, 
                                                           f_c, f_c_type, 
                                                           recombination, recombination, recombination_inv, recombination_inv, 
                                                           m_r, 
                                                           a_r, a_r_i_f, a_r_i_g, 
                                                           a_r_g, a_r_f)
                    for level in range(1, gen_level+1):
                        p_f_ave_1 = np.mean(all_p_f_1[:level*100], axis=0)
                        a_f_ave_1 = np.mean(all_a_f_1[:level*100], axis=0)
                        
                        p_f_ave_2 = np.mean(all_p_f_2[:level*100], axis=0)
                        a_f_ave_2 = np.mean(all_a_f_2[:level*100], axis=0)

                        for a_i, a in enumerate(all_alleles):
                            data = [[m_r, r_d, p_f_ave_1[a_i], 'phenotype', level*100, a, 1],
                                    [m_r, r_d, a_f_ave_1[a_i], 'allele', level*100, a, 1], 
                                    [m_r, r_d, p_f_ave_2[a_i], 'phenotype', level*100, a, 2],
                                    [m_r, r_d, a_f_ave_2[a_i], 'allele', level*100, a, 2]]
                            
                            df = df.append(pd.DataFrame(data, columns = cols))
                            
                            phenotypes_matrix[0, level-1, a_i, r_d_i, m_r_i] = p_f_ave_1[a_i]
                            alleles_matrix[0, level-1, a_i, r_d_i, m_r_i] = a_f_ave_1[a_i]
                            
                            phenotypes_matrix[1, level-1, a_i, r_d_i, m_r_i] = p_f_ave_2[a_i]
                            alleles_matrix[1, level-1, a_i, r_d_i, m_r_i] = a_f_ave_2[a_i]
        
        sheet_name = sheet + '_2pop'
        
        for pop in range(num_pops):

            for level in range(1, gen_level+1):
                for a_i, a in enumerate(all_alleles):
                    df_p_1 = pd.DataFrame(phenotypes_matrix[pop, level-1, a_i])
                    df_a_1 = pd.DataFrame(alleles_matrix[pop, level-1, a_i])

                    append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                       df_p_1, pop+1, 
                                       gene_drive_name, '\'Phenotype\' frequency', 
                                       i_f_list, i_g, 
                                       f_c_c_list, f_c_r_list, 
                                       r_d_list, 
                                       m_r_list, 
                                       a_r, a_r_i_f, a_r_i_g, 
                                       a_r_g, a_r_f, 
                                       1, a, level*100)
                    append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                       df_a_1, pop+1, 
                                       gene_drive_name, 'Allele frequency', 
                                       i_f_list, i_g, 
                                       f_c_c_list, f_c_r_list, 
                                       r_d_list, 
                                       m_r_list, 
                                       a_r, a_r_i_f, a_r_i_g, 
                                       a_r_g, a_r_f, 
                                       1, a, level*100)
        
        # Plot Heatmap(s)
        temp_plot = heatmap_plots(gene_drive_name, h_type, df, num_pops, plot_alleles,  num_gen_avg, figure_width, figure_height)

        if num_pops == 1:
            data_plot = temp_plot.cols(1)

        elif num_pops == 2:
            data_plot = temp_plot.cols(2)

        else:
            data_plot = temp_plot.cols(3)
        
        return(df, data_plot)
    
    elif num_pops == 3:
        
        if h_type == 'FCvIF':
            
            phenotypes_matrix = np.array([[[[[None]*(f_c_steps)]*(i_f_steps)]*len(all_alleles)]*gen_level]*num_pops)
            alleles_matrix = np.array([[[[[None]*(f_c_steps)]*(i_f_steps)]*len(all_alleles)]*gen_level]*num_pops)
            
            cols = ['Fitness Cost', 
                    'Introduction Frequency', 
                    'Averaged Frequency', 
                    'Frequency Type', 
                    'Number of Generations Data Averaged Over', 
                    'Allele', 
                    'Population']
            
            recombination = list(np.zeros(num_loci-1))
            recombination[0] = 1-r_d_list[0]/50

            recombination_inv = list(np.zeros(num_loci-1))
            recombination_inv[0] = 1-r_d_inv_list[0]/50

            male_cm, female_cm = matrix_evaluator(male_cross_matrix, female_cross_matrix, d_a, 
                                                  d_a_males_string, d_a_females_string, 
                                                  d_a_males_rates_string, d_a_females_rates_string, 
                                                  num_loci, num_male_geno, num_female_geno, 
                                                  drive_activites_total_counter, 
                                                  recombination, recombination, recombination_inv, recombination_inv)

            m_eval = [male_cm, female_cm]
            
            m_r = m_r_list[0]
            
            for i_f_i, i_f in enumerate(i_f_list):
                i_f_inv = i_f_inv_list[i_f_i]
                for f_c_ind, f_c_c in enumerate(f_c_c_list):
                    f_c_r = f_c_r_list[f_c_ind]
                    f_c_i = f_c_r_list[f_c_ind]
                    f_c_male_alleles = np.zeros(len(male_alleles))
                    f_c_male_alleles[0] = f_c_c
                    f_c_female_alleles = np.zeros(len(female_alleles))
                    f_c_female_alleles[0] = f_c_c

                    if 'I' in all_alleles:
                        f_c_male_alleles[all_alleles.index('I')] = f_c_i
                        f_c_female_alleles[all_alleles.index('I')] = f_c_i

                    if 'R' in all_alleles:
                        f_c_male_alleles[all_alleles.index('R')] = f_c_r
                        f_c_female_alleles[all_alleles.index('R')] = f_c_r

                    f_c = [f_c_male_alleles, f_c_female_alleles]
                    
                    all_p_f_1, all_a_f_1, \
                    all_p_f_2, all_a_f_2, \
                    all_p_f_3, all_a_f_3 = drive_data_3pop(master_list, d_a, lethal, rescue, 
                                                           m_eval, num_gens, 
                                                           i_f, i_g, 
                                                           i_f_inv, i_inv_g, 
                                                           f_c, f_c_type, 
                                                           recombination, recombination, recombination_inv, recombination_inv, 
                                                           m_r, 
                                                           a_r, a_r_i_f, a_r_i_g, 
                                                           a_r_g, a_r_f)
                    for level in range(1, gen_level+1):
                        p_f_ave_1 = np.mean(all_p_f_1[:level*100], axis=0)
                        a_f_ave_1 = np.mean(all_a_f_1[:level*100], axis=0)
                        
                        p_f_ave_2 = np.mean(all_p_f_2[:level*100], axis=0)
                        a_f_ave_2 = np.mean(all_a_f_2[:level*100], axis=0)
                        
                        p_f_ave_3 = np.mean(all_p_f_3[:level*100], axis=0)
                        a_f_ave_3 = np.mean(all_a_f_3[:level*100], axis=0)

                        for a_i, a in enumerate(all_alleles):
                            data = [[f_c_c, np.sum(i_f), p_f_ave_1[a_i], 'phenotype', level*100, a, 1],
                                    [f_c_c, np.sum(i_f), a_f_ave_1[a_i], 'allele', level*100, a, 1], 
                                    [f_c_c, np.sum(i_f), p_f_ave_2[a_i], 'phenotype', level*100, a, 2],
                                    [f_c_c, np.sum(i_f), a_f_ave_2[a_i], 'allele', level*100, a, 2], 
                                    [f_c_c, np.sum(i_f), p_f_ave_3[a_i], 'phenotype', level*100, a, 3],
                                    [f_c_c, np.sum(i_f), a_f_ave_3[a_i], 'allele', level*100, a, 3]]
                            
                            df = df.append(pd.DataFrame(data, columns = cols))
                            
                            phenotypes_matrix[0, level-1, a_i, i_f_i, f_c_ind] = p_f_ave_1[a_i]
                            alleles_matrix[0, level-1, a_i, i_f_i, f_c_ind] = a_f_ave_1[a_i]
                            
                            phenotypes_matrix[1, level-1, a_i, i_f_i, f_c_ind] = p_f_ave_2[a_i]
                            alleles_matrix[1, level-1, a_i, i_f_i, f_c_ind] = a_f_ave_2[a_i]
                            
                            phenotypes_matrix[2, level-1, a_i, i_f_i, f_c_ind] = p_f_ave_3[a_i]
                            alleles_matrix[2, level-1, a_i, i_f_i, f_c_ind] = a_f_ave_3[a_i]

        elif h_type == 'RDvIF':
            
            phenotypes_matrix = np.array([[[[[None]*(r_d_steps)]*(i_f_steps)]*len(all_alleles)]*gen_level]*num_pops)
            alleles_matrix = np.array([[[[[None]*(r_d_steps)]*(i_f_steps)]*len(all_alleles)]*gen_level]*num_pops)
            
            cols = ['Recombination Distance', 
                    'Introduction Frequency', 
                    'Averaged Frequency', 
                    'Frequency Type', 
                    'Number of Generations Data Averaged Over', 
                    'Allele', 
                    'Population']
            
            f_c_c = f_c_c_list[0]
            f_c_r = f_c_r_list[0]
            f_c_i = f_c_i_list[0]
            
            f_c_male_alleles = np.zeros(len(male_alleles))
            f_c_male_alleles[0] = f_c_c
            f_c_female_alleles = np.zeros(len(female_alleles))
            f_c_female_alleles[0] = f_c_c

            if 'I' in all_alleles:
                f_c_male_alleles[all_alleles.index('I')] = f_c_i
                f_c_female_alleles[all_alleles.index('I')] = f_c_i

            if 'R' in all_alleles:
                f_c_male_alleles[all_alleles.index('R')] = f_c_r
                f_c_female_alleles[all_alleles.index('R')] = f_c_r

            f_c = [f_c_male_alleles, f_c_female_alleles]
            
            m_r = m_r_list[0]
            
            for i_f_i, i_f in enumerate(i_f_list):
                i_f_inv = i_f_inv_list[i_f_i]
                for r_d_i, r_d in enumerate(r_d_list):
                    
                    recombination = list(np.zeros(num_loci-1))
                    recombination[0] = 1-r_d/50

                    recombination_inv = list(np.zeros(num_loci-1))
                    recombination_inv[0] = 1-r_d_inv_list[r_d_i]/50
                    
                    all_p_f_1, all_a_f_1, \
                    all_p_f_2, all_a_f_2, \
                    all_p_f_3, all_a_f_3 = drive_data_3pop(master_list, d_a, lethal, rescue, 
                                                           0, num_gens, 
                                                           i_f, i_g, 
                                                           i_f_inv, i_inv_g, 
                                                           f_c, f_c_type, 
                                                           recombination, recombination, recombination_inv, recombination_inv, 
                                                           m_r, 
                                                           a_r, a_r_i_f, a_r_i_g, 
                                                           a_r_g, a_r_f)
                    for level in range(1, gen_level+1):
                        p_f_ave_1 = np.mean(all_p_f_1[:level*100], axis=0)
                        a_f_ave_1 = np.mean(all_a_f_1[:level*100], axis=0)
                        
                        p_f_ave_2 = np.mean(all_p_f_2[:level*100], axis=0)
                        a_f_ave_2 = np.mean(all_a_f_2[:level*100], axis=0)
                        
                        p_f_ave_3 = np.mean(all_p_f_3[:level*100], axis=0)
                        a_f_ave_3 = np.mean(all_a_f_3[:level*100], axis=0)

                        for a_i, a in enumerate(all_alleles):
                            data = [[r_d, np.sum(i_f), p_f_ave_1[a_i], 'phenotype', level*100, a, 1],
                                    [r_d, np.sum(i_f), a_f_ave_1[a_i], 'allele', level*100, a, 1], 
                                    [r_d, np.sum(i_f), p_f_ave_2[a_i], 'phenotype', level*100, a, 2],
                                    [r_d, np.sum(i_f), a_f_ave_2[a_i], 'allele', level*100, a, 2], 
                                    [r_d, np.sum(i_f), p_f_ave_3[a_i], 'phenotype', level*100, a, 3],
                                    [r_d, np.sum(i_f), a_f_ave_3[a_i], 'allele', level*100, a, 3]]
                            
                            df = df.append(pd.DataFrame(data, columns = cols))
                            
                            phenotypes_matrix[0, level-1, a_i, i_f_i, r_d_i] = p_f_ave_1[a_i]
                            alleles_matrix[0, level-1, a_i, i_f_i, r_d_i] = a_f_ave_1[a_i]
                            
                            phenotypes_matrix[1, level-1, a_i, i_f_i, r_d_i] = p_f_ave_2[a_i]
                            alleles_matrix[1, level-1, a_i, i_f_i, r_d_i] = a_f_ave_2[a_i]
                            
                            phenotypes_matrix[2, level-1, a_i, i_f_i, r_d_i] = p_f_ave_3[a_i]
                            alleles_matrix[2, level-1, a_i, i_f_i, r_d_i] = a_f_ave_3[a_i]

        elif h_type == 'FCvRD':
            
            phenotypes_matrix = np.array([[[[[None]*(f_c_steps)]*(r_d_steps)]*len(all_alleles)]*gen_level]*num_pops)
            alleles_matrix = np.array([[[[[None]*(f_c_steps)]*(r_d_steps)]*len(all_alleles)]*gen_level]*num_pops)
            
            cols = ['Fitness Cost', 
                    'Recombination Distance', 
                    'Averaged Frequency', 
                    'Frequency Type', 
                    'Number of Generations Data Averaged Over', 
                    'Allele', 
                    'Population']
            
            i_f = i_f_list[0]
            i_f_inv = i_f_inv_list[0]
            
            m_r = m_r_list[0]
            
            for r_d_i, r_d in enumerate(r_d_list):
                for f_c_ind, f_c_c in enumerate(f_c_c_list):
                    f_c_r = f_c_r_list[f_c_ind]
                    f_c_i = f_c_r_list[f_c_ind]
                    f_c_male_alleles = np.zeros(len(male_alleles))
                    f_c_male_alleles[0] = f_c_c
                    f_c_female_alleles = np.zeros(len(female_alleles))
                    f_c_female_alleles[0] = f_c_c

                    if 'I' in all_alleles:
                        f_c_male_alleles[all_alleles.index('I')] = f_c_i
                        f_c_female_alleles[all_alleles.index('I')] = f_c_i

                    if 'R' in all_alleles:
                        f_c_male_alleles[all_alleles.index('R')] = f_c_r
                        f_c_female_alleles[all_alleles.index('R')] = f_c_r

                    f_c = [f_c_male_alleles, f_c_female_alleles]
                    
                    recombination = list(np.zeros(num_loci-1))
                    recombination[0] = 1-r_d/50

                    recombination_inv = list(np.zeros(num_loci-1))
                    recombination_inv[0] = 1-r_d_inv_list[r_d_i]/50
                    
                    all_p_f_1, all_a_f_1, \
                    all_p_f_2, all_a_f_2, \
                    all_p_f_3, all_a_f_3 = drive_data_3pop(master_list, d_a, lethal, rescue, 
                                                           0, num_gens, 
                                                           i_f, i_g, 
                                                           i_f_inv, i_inv_g, 
                                                           f_c, f_c_type, 
                                                           recombination, recombination, recombination_inv, recombination_inv, 
                                                           m_r, 
                                                           a_r, a_r_i_f, a_r_i_g, 
                                                           a_r_g, a_r_f)
                    for level in range(1, gen_level+1):
                        p_f_ave_1 = np.mean(all_p_f_1[:level*100], axis=0)
                        a_f_ave_1 = np.mean(all_a_f_1[:level*100], axis=0)
                        
                        p_f_ave_2 = np.mean(all_p_f_2[:level*100], axis=0)
                        a_f_ave_2 = np.mean(all_a_f_2[:level*100], axis=0)
                        
                        p_f_ave_3 = np.mean(all_p_f_3[:level*100], axis=0)
                        a_f_ave_3 = np.mean(all_a_f_3[:level*100], axis=0)

                        for a_i, a in enumerate(all_alleles):
                            data = [[f_c_c, r_d, p_f_ave_1[a_i], 'phenotype', level*100, a, 1],
                                    [f_c_c, r_d, a_f_ave_1[a_i], 'allele', level*100, a, 1], 
                                    [f_c_c, r_d, p_f_ave_2[a_i], 'phenotype', level*100, a, 2],
                                    [f_c_c, r_d, a_f_ave_2[a_i], 'allele', level*100, a, 2], 
                                    [f_c_c, r_d, p_f_ave_3[a_i], 'phenotype', level*100, a, 3],
                                    [f_c_c, r_d, a_f_ave_3[a_i], 'allele', level*100, a, 3]]
                            
                            df = df.append(pd.DataFrame(data, columns = cols))
                            
                            phenotypes_matrix[0, level-1, a_i, r_d_i, f_c_ind] = p_f_ave_1[a_i]
                            alleles_matrix[0, level-1, a_i, r_d_i, f_c_ind] = a_f_ave_1[a_i]
                            
                            phenotypes_matrix[1, level-1, a_i, r_d_i, f_c_ind] = p_f_ave_2[a_i]
                            alleles_matrix[1, level-1, a_i, r_d_i, f_c_ind] = a_f_ave_2[a_i]
                            
                            phenotypes_matrix[2, level-1, a_i, r_d_i, f_c_ind] = p_f_ave_3[a_i]
                            alleles_matrix[2, level-1, a_i, r_d_i, f_c_ind] = a_f_ave_3[a_i]

        elif h_type == 'MRvRD':
            
            phenotypes_matrix = np.array([[[[[None]*(m_r_steps)]*(r_d_steps)]*len(all_alleles)]*gen_level]*num_pops)
            alleles_matrix = np.array([[[[[None]*(m_r_steps)]*(r_d_steps)]*len(all_alleles)]*gen_level]*num_pops)
            
            cols = ['Migration Rate', 
                    'Recombination Distance', 
                    'Averaged Frequency', 
                    'Frequency Type', 
                    'Number of Generations Data Averaged Over', 
                    'Allele', 
                    'Population']
            
            i_f = i_f_list[0]
            i_f_inv = i_f_inv_list[0]
            
            f_c_c = f_c_c_list[0]
            f_c_r = f_c_r_list[0]
            f_c_i = f_c_i_list[0]
            
            f_c_male_alleles = np.zeros(len(male_alleles))
            f_c_male_alleles[0] = f_c_c
            f_c_female_alleles = np.zeros(len(female_alleles))
            f_c_female_alleles[0] = f_c_c

            if 'I' in all_alleles:
                f_c_male_alleles[all_alleles.index('I')] = f_c_i
                f_c_female_alleles[all_alleles.index('I')] = f_c_i

            if 'R' in all_alleles:
                f_c_male_alleles[all_alleles.index('R')] = f_c_r
                f_c_female_alleles[all_alleles.index('R')] = f_c_r

            f_c = [f_c_male_alleles, f_c_female_alleles]
            
            m_r = m_r_list[0]
            
            for r_d_i, r_d in enumerate(r_d_list):
                for m_r_i, m_r in enumerate(m_r_list):
                    
                    recombination = list(np.zeros(num_loci-1))
                    recombination[0] = 1-r_d/50

                    recombination_inv = list(np.zeros(num_loci-1))
                    recombination_inv[0] = 1-r_d_inv_list[r_d_i]/50
                    
                    all_p_f_1, all_a_f_1, \
                    all_p_f_2, all_a_f_2, \
                    all_p_f_3, all_a_f_3 = drive_data_3pop(master_list, d_a, lethal, rescue, 
                                                           0, num_gens, 
                                                           i_f, i_g, 
                                                           i_f_inv, i_inv_g, 
                                                           f_c, f_c_type, 
                                                           recombination, recombination, recombination_inv, recombination_inv, 
                                                           m_r, 
                                                           a_r, a_r_i_f, a_r_i_g, 
                                                           a_r_g, a_r_f)
                    for level in range(1, gen_level+1):
                        p_f_ave_1 = np.mean(all_p_f_1[:level*100], axis=0)
                        a_f_ave_1 = np.mean(all_a_f_1[:level*100], axis=0)
                        
                        p_f_ave_2 = np.mean(all_p_f_2[:level*100], axis=0)
                        a_f_ave_2 = np.mean(all_a_f_2[:level*100], axis=0)
                        
                        p_f_ave_3 = np.mean(all_p_f_3[:level*100], axis=0)
                        a_f_ave_3 = np.mean(all_a_f_3[:level*100], axis=0)

                        for a_i, a in enumerate(all_alleles):
                            data = [[m_r, r_d, p_f_ave_1[a_i], 'phenotype', level*100, a, 1],
                                    [m_r, r_d, a_f_ave_1[a_i], 'allele', level*100, a, 1], 
                                    [m_r, r_d, p_f_ave_2[a_i], 'phenotype', level*100, a, 2],
                                    [m_r, r_d, a_f_ave_2[a_i], 'allele', level*100, a, 2], 
                                    [m_r, r_d, p_f_ave_3[a_i], 'phenotype', level*100, a, 3],
                                    [m_r, r_d, a_f_ave_3[a_i], 'allele', level*100, a, 3]]
                            
                            df = df.append(pd.DataFrame(data, columns = cols))
                            
                            phenotypes_matrix[0, level-1, a_i, r_d_i, m_r_i] = p_f_ave_1[a_i]
                            alleles_matrix[0, level-1, a_i, r_d_i, m_r_i] = a_f_ave_1[a_i]
                            
                            phenotypes_matrix[1, level-1, a_i, r_d_i, m_r_i] = p_f_ave_2[a_i]
                            alleles_matrix[1, level-1, a_i, r_d_i, m_r_i] = a_f_ave_2[a_i]
                            
                            phenotypes_matrix[2, level-1, a_i, r_d_i, m_r_i] = p_f_ave_3[a_i]
                            alleles_matrix[2, level-1, a_i, r_d_i, m_r_i] = a_f_ave_3[a_i]
                            
        sheet_name = sheet + '_3pop'
        
        for pop in range(num_pops):

            for level in range(1, gen_level+1):
                for a_i, a in enumerate(all_alleles):
                    df_p_1 = pd.DataFrame(phenotypes_matrix[pop, level-1, a_i])
                    df_a_1 = pd.DataFrame(alleles_matrix[pop, level-1, a_i])

                    append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                       df_p_1, pop+1, 
                                       gene_drive_name, '\'Phenotype\' frequency', 
                                       i_f_list, i_g, 
                                       f_c_c_list, f_c_r_list, 
                                       r_d_list, 
                                       m_r_list, 
                                       a_r, a_r_i_f, a_r_i_g, 
                                       a_r_g, a_r_f, 
                                       1, a, level*100)
                    append_df_to_excel('/Users/Tobin/Desktop/' + file + '.xlsx', sheet, 
                                       df_a_1, pop+1, 
                                       gene_drive_name, 'Allele frequency', 
                                       i_f_list, i_g, 
                                       f_c_c_list, f_c_r_list, 
                                       r_d_list, 
                                       m_r_list, 
                                       a_r, a_r_i_f, a_r_i_g, 
                                       a_r_g, a_r_f, 
                                       1, a, level*100)
        
        # Plot Heatmap(s)
        temp_plot = heatmap_plots(gene_drive_name, h_type, df, num_pops, plot_alleles,  num_gen_avg, figure_width, figure_height)

        if num_pops == 1:
            data_plot = temp_plot.cols(1)

        elif num_pops == 2:
            data_plot = temp_plot.cols(2)

        else:
            data_plot = temp_plot.cols(3)
        
        return(df, data_plot)
        
    else:
        raise RuntimeError(f'Unsupported heatmap pairing, check \'steps\' inputs.')


def heatmap_plots(gene_drive_name, h_type, df, num_pops, plot_alleles, num_gen_avg, fig_width, fig_height):
    """
    heatmap_plots handles plotting for the heatmap function.
    """
    
    plot_list = []
    
    if h_type == 'FCvIF':
        for allele in plot_alleles:
            for pop in range(1, num_pops+1):
            
                plot = hv.HeatMap(
                    data=df[(df['Allele'] == allele) \
                            & (df['Frequency Type'] == 'phenotype') \
                            & (df['Number of Generations Data Averaged Over'] == num_gen_avg) \
                            & (df['Population'] == pop)],
                    kdims=['Fitness Cost', 'Introduction Frequency'],
                    vdims=['Averaged Frequency'],
                ).opts(
                    title=str(gene_drive_name) + ' Frequency of ' + str(allele) + ' in population ' + str(pop),
                    tools=['hover'],
                    width=fig_width,
                    height=fig_height
                )

                plot_list.append(plot)
                
    elif h_type == 'RDvIF':
        for allele in plot_alleles:
            for pop in range(1, num_pops+1):
            
                plot = hv.HeatMap(
                    data=df[(df['Allele'] == allele) \
                            & (df['Frequency Type'] == 'phenotype') \
                            & (df['Number of Generations Data Averaged Over'] == num_gen_avg) \
                            & (df['Population'] == pop)],
                    kdims=['Recombination Distance', 'Introduction Frequency'],
                    vdims=['Averaged Frequency'],
                ).opts(
                    title=str(gene_drive_name) + ' Frequency of ' + str(allele) + ' in population ' + str(pop),
                    tools=['hover'],
                    width=fig_width,
                    height=fig_height
                )

                plot_list.append(plot)
                
    elif h_type == 'FCvRD':
        for allele in plot_alleles:
            for pop in range(1, num_pops+1):
            
                plot = hv.HeatMap(
                    data=df[(df['Allele'] == allele) \
                            & (df['Frequency Type'] == 'phenotype') \
                            & (df['Number of Generations Data Averaged Over'] == num_gen_avg) \
                            & (df['Population'] == pop)],
                    kdims=['Fitness Cost', 'Recombination Distance'],
                    vdims=['Averaged Frequency'],
                ).opts(
                    title=str(gene_drive_name) + ' Frequency of ' + str(allele) + ' in population ' + str(pop),
                    tools=['hover'],
                    width=fig_width,
                    height=fig_height
                )
                plot_list.append(plot)
                
    elif h_type == 'MRvRD':
        for allele in plot_alleles:
            for pop in range(1, num_pops+1):
            
                plot = hv.HeatMap(
                    data=df[(df['Allele'] == allele) \
                            & (df['Frequency Type'] == 'phenotype') \
                            & (df['Number of Generations Data Averaged Over'] == num_gen_avg) \
                            & (df['Population'] == pop)],
                    kdims=['Migration Rate', 'Recombination Distance'],
                    vdims=['Averaged Frequency'],
                ).opts(
                    title=str(gene_drive_name) + ' Frequency of ' + str(allele) + ' in population ' + str(pop),
                    tools=['hover'],
                    width=fig_width,
                    height=fig_height
                )

                plot_list.append(plot)
    
    else:
        raise RuntimeError(f'Unsupported heatmap pairing, check \'steps\' inputs.')
    
    data_plot = plot_list[0]
    
    if len(plot_list) >= 2:
        for p in range(1, len(plot_list)):
            data_plot += plot_list[p]

    return(data_plot)

# Reference for the bebi103 package
# @misc{#10.22002/D1.1615,
#   doi = {10.22002/D1.1615},
#   url = {https://doi.org/10.22002/D1.1615},
#   author = {Bois, Justin S.},
#   keywords = {Github},
#   title = {justinbois/bebi103: Version 0.1.0},
#   publisher = {CaltechDATA},
#   year = {2020}
# }